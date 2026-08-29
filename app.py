import os,sys,shutil,hashlib,datetime,uuid,sqlite3,threading
from functools import wraps
from io import BytesIO
from flask import Flask,jsonify,request,session,send_from_directory,send_file,Response
import openpyxl
from openpyxl.styles import Font,Alignment,PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

app = Flask(__name__)
app.secret_key = "xtsgj_contract_manager_secret_key_2026"
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATA_DIR = os.path.join(BASE_DIR,"data")
BACKUP_DIR = os.path.join(BASE_DIR,"backups")
STATIC_DIR = os.path.join(BASE_DIR,"static")
EXCEL_PATH = os.path.join(DATA_DIR,"\u5408\u540c\u53f0\u8d26.xlsx")
DB_PATH = os.path.join(DATA_DIR,"合同台账.db")
ATTACH_DIR = os.path.join(DATA_DIR,"attachments")
ATTACH_SHEET = "\u5408\u540c\u9644\u4ef6"
ATTACH_HEADERS = ["\u9644\u4ef6ID","\u5408\u540c\u7f16\u53f7","\u9644\u4ef6\u7c7b\u578b","\u663e\u793a\u540d\u79f0","\u5b58\u50a8\u6587\u4ef6\u540d","\u6587\u4ef6\u5927\u5c0f","\u4e0a\u4f20\u4eba","\u4e0a\u4f20\u65f6\u95f4"]
ATTACH_TYPES = {"decision":"\u51b3\u7b56\u4f9d\u636e\u9644\u4ef6","contract":"\u5408\u540c\u9644\u4ef6","acceptance":"\u9a8c\u6536\u7ed3\u7b97\u9644\u4ef6"}
ATTACH_ALLOWED_EXT = {".jpg",".jpeg",".png",".gif",".webp",".bmp",".pdf"}
DEFAULT_ATTACH_MAX_MB = 20
SALT = "xtsgj_2024_contract"

for d in [DATA_DIR,BACKUP_DIR,STATIC_DIR,ATTACH_DIR,os.path.join(BASE_DIR,"templates")]:
    os.makedirs(d,exist_ok=True)

CONTRACT_COLUMNS = [
    "\u7f16\u53f7","\u9879\u76ee\u5206\u7c7b","\u9879\u76ee\u540d\u79f0","\u51b3\u7b56\u4f9d\u636e",
    "\u51b3\u7b56\u7684\u62db\u6807\u63a7\u5236\u4ef7","\u62db\u6807\u63a7\u5236\u4ef7\u51b3\u7b56\u7684\u4f9d\u636e",
    "\u91c7\u8d2d\u65b9\u5f0f","\u4e2d\u6807\u5355\u4f4d","\u5b9e\u9645\u5b9e\u65bd\u5355\u4f4d",
    "\u5408\u540c\u91d1\u989d\uff08\u4e07\u5143\uff09","\u8d44\u91d1\u6765\u6e90","\u5408\u540c\u4e3b\u8981\u5185\u5bb9",
    "\u91c7\u8d2d\u65f6\u95f4","\u5f00\u5de5\u65f6\u95f4","\u5b8c\u5de5\u65f6\u95f4",
    "\u7ae3\u5de5\u9a8c\u6536\u65f6\u95f4","\u73b0\u573a\u8d1f\u8d23\u4eba",
    "\u662f\u5426\u6709\u7b7e\u8bc1\u6216\u53d8\u66f4","\u7b7e\u8bc1\u6216\u53d8\u66f4\u60c5\u51b5",
    "\u662f\u5426\u7ed3\u7b97","\u7ed3\u7b97\u91d1\u989d","\u5df2\u652f\u4ed8","\u672a\u652f\u4ed8",
    "\u671f\u9650","\u7ecf\u529e\u4eba","\u5907\u6ce8",
    "\u6240\u5c5e\u90e8\u95e8","\u4e0b\u6b21\u62db\u91c7\u65f6\u95f4","\u63d0\u9192\u63d0\u524d\u5929\u6570",
]

COL_IDX = {n:i for i,n in enumerate(CONTRACT_COLUMNS)}

# ──── SQLite I/O ────
db_lock = threading.Lock()

def _conn():
    con = sqlite3.connect(DB_PATH, timeout=15)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA busy_timeout=15000")
    con.execute("PRAGMA foreign_keys=ON")
    return con

def db_query(sql, params=()):
    con = _conn()
    try:
        cur = con.execute(sql, params)
        return [{k: r[k] for k in r.keys()} for r in cur.fetchall()]
    finally:
        con.close()

def db_exec(sql, params=()):
    with db_lock:
        _backup()
        con = _conn()
        try:
            cur = con.execute(sql, params)
            con.commit()
            return cur
        except Exception:
            con.rollback()
            raise
        finally:
            con.close()

def _backup():
    if not os.path.exists(DB_PATH): return
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    dest_path = os.path.join(BACKUP_DIR, f"合同台账_{ts}.db")
    src = sqlite3.connect(DB_PATH, timeout=15)
    dst = sqlite3.connect(dest_path)
    try:
        src.backup(dst)
    finally:
        dst.close(); src.close()
    cutoff = datetime.datetime.now() - datetime.timedelta(days=30)
    for fn in os.listdir(BACKUP_DIR):
        fp = os.path.join(BACKUP_DIR, fn)
        if fn.endswith(".db") and datetime.datetime.fromtimestamp(os.path.getmtime(fp)) < cutoff:
            try: os.remove(fp)
            except: pass

_SCHEMA = """
CREATE TABLE IF NOT EXISTS contracts (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  "编号" TEXT NOT NULL, "项目分类" TEXT NOT NULL, "项目名称" TEXT NOT NULL, "决策依据" TEXT,
  "决策的招标控制价" REAL CHECK ("决策的招标控制价" IS NULL OR CAST("决策的招标控制价" AS REAL) >= 0),
  "招标控制价决策的依据" TEXT, "采购方式" TEXT NOT NULL, "中标单位" TEXT, "实际实施单位" TEXT,
  "合同金额（万元）" REAL NOT NULL CHECK ("合同金额（万元）" IS NULL OR CAST("合同金额（万元）" AS REAL) >= 0),
  "资金来源" TEXT NOT NULL, "合同主要内容" TEXT NOT NULL,
  "采购时间" TEXT CHECK ("采购时间" IS NULL OR "采购时间" = '' OR "采购时间" GLOB '[0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]'),
  "开工时间" TEXT CHECK ("开工时间" IS NULL OR "开工时间" = '' OR "开工时间" GLOB '[0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]'),
  "完工时间" TEXT CHECK ("完工时间" IS NULL OR "完工时间" = '' OR "完工时间" GLOB '[0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]'),
  "竣工验收时间" TEXT CHECK ("竣工验收时间" IS NULL OR "竣工验收时间" = '' OR "竣工验收时间" GLOB '[0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]'),
  "现场负责人" TEXT,
  "是否有签证或变更" TEXT, "签证或变更情况" TEXT,
  "是否结算" TEXT, "结算金额" REAL CHECK ("结算金额" IS NULL OR CAST("结算金额" AS REAL) >= 0),
  "已支付" REAL CHECK ("已支付" IS NULL OR CAST("已支付" AS REAL) >= 0),
  "未支付" REAL CHECK ("未支付" IS NULL OR CAST("未支付" AS REAL) >= 0),
  "期限" TEXT, "经办人" TEXT NOT NULL, "备注" TEXT,
  "所属部门" TEXT NOT NULL,
  "下次招采时间" TEXT CHECK ("下次招采时间" IS NULL OR "下次招采时间" = '' OR "下次招采时间" GLOB '[0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]'),
  "提醒提前天数" INTEGER,
  "是否归档" TEXT, "归档人" TEXT, "归档时间" TEXT,
  "所属单位" TEXT, "所属科室" TEXT,
  -- A-05 枚举 CHECK（项目分类/采购方式/是否结算/资金来源/是否归档）延后至 A-08 清洗完成后启用
  FOREIGN KEY ("所属单位","所属科室") REFERENCES departments("单位名称","科室名称") ON DELETE RESTRICT ON UPDATE CASCADE
);
CREATE TABLE IF NOT EXISTS users (
  "用户名" TEXT PRIMARY KEY, "密码" TEXT, "姓名" TEXT, "所属部门" TEXT, "角色" TEXT, "分管部门" TEXT,
  "所属单位" TEXT, "所属科室" TEXT,
  FOREIGN KEY ("所属单位","所属科室") REFERENCES departments("单位名称","科室名称") ON DELETE RESTRICT ON UPDATE CASCADE
);
CREATE TABLE IF NOT EXISTS departments (
  "单位名称" TEXT, "科室名称" TEXT
);
CREATE TABLE IF NOT EXISTS settings (
  "配置项" TEXT PRIMARY KEY, "配置值" TEXT
);
CREATE TABLE IF NOT EXISTS attachments (
  "附件ID" TEXT PRIMARY KEY, "合同编号" TEXT, "附件类型" TEXT, "显示名称" TEXT,
  "存储文件名" TEXT, "文件大小" TEXT, "上传人" TEXT, "上传时间" TEXT,
  FOREIGN KEY ("合同编号") REFERENCES contracts("编号") ON DELETE RESTRICT ON UPDATE RESTRICT
);
CREATE TABLE IF NOT EXISTS operation_logs (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  "时间" TEXT, "用户名" TEXT, "姓名" TEXT, "角色" TEXT, "IP" TEXT,
  "模块" TEXT, "操作类型" TEXT, "对象" TEXT, "详情" TEXT, "结果" TEXT
);
"""

# ─── 数据库结构版本化迁移（schema_version 记录已应用版本；迁移函数须保持幂等）───
def _mig_v1_contract_extra_cols(con):
    """v1：contracts 追加历史字段列（资金来源/是否归档/归档人/归档时间），幂等（存在即跳过）。"""
    cols = [r[1] for r in con.execute("PRAGMA table_info(contracts)").fetchall()]
    for cname, ctype in (("资金来源","TEXT"),("是否归档","TEXT"),("归档人","TEXT"),("归档时间","TEXT")):
        if cname not in cols:
            con.execute('ALTER TABLE contracts ADD COLUMN "' + cname + '" ' + ctype)

def _mig_v2_contract_rename_cols(con):
    """v2：contracts 历史列改名（经办人/分管领导→经办人、甲方现场负责人→现场负责人），幂等。"""
    cols = [r[1] for r in con.execute("PRAGMA table_info(contracts)").fetchall()]
    if "经办人/分管领导" in cols and "经办人" not in cols:
        con.execute('ALTER TABLE contracts RENAME COLUMN "经办人/分管领导" TO "经办人"')
    if "甲方现场负责人" in cols and "现场负责人" not in cols:
        con.execute('ALTER TABLE contracts RENAME COLUMN "甲方现场负责人" TO "现场负责人"')

def _mig_v3_users_supervise_dept(con):
    """v3：users 追加 分管部门 字段，幂等。"""
    ucols = [r[1] for r in con.execute("PRAGMA table_info(users)").fetchall()]
    if "分管部门" not in ucols:
        con.execute('ALTER TABLE users ADD COLUMN "分管部门" TEXT')

MIGRATIONS = [
    (1, "contracts 追加历史字段列：资金来源/是否归档/归档人/归档时间", _mig_v1_contract_extra_cols),
    (2, "contracts 列改名：经办人/分管领导→经办人、甲方现场负责人→现场负责人", _mig_v2_contract_rename_cols),
    (3, "users 追加 分管部门 字段", _mig_v3_users_supervise_dept),
]

def _run_versioned_migrations(con):
    """按版本号顺序执行未应用的数据库结构迁移。

    - 建表：schema_version(version INTEGER PRIMARY KEY, applied_at TEXT, description TEXT)。
    - 仅执行 version > 当前最大已应用版本的迁移；每个迁移独立事务，成功后写入版本记录。
    - 失败：回滚事务、不记录版本、向上抛出错误（下次启动重试）。
    - 迁移函数本身保持幂等（含存在性判断），删除版本行可强制重跑对应版本。
    - 新增结构变更须向 MIGRATIONS 追加新版本号，禁止修改已执行的历史版本。
    """
    con.execute("CREATE TABLE IF NOT EXISTS schema_version (version INTEGER PRIMARY KEY, applied_at TEXT, description TEXT)")
    applied = set(r[0] for r in con.execute("SELECT version FROM schema_version").fetchall())
    for ver, desc, fn in MIGRATIONS:
        if ver in applied:
            continue
        con.execute("BEGIN")
        try:
            fn(con)
            ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            con.execute("INSERT INTO schema_version (version, applied_at, description) VALUES (?,?,?)", (ver, ts, desc))
            con.execute("COMMIT")
        except Exception:
            con.execute("ROLLBACK")
            raise
        print(f"[MIGRATE] schema_version {ver}: {desc}", file=sys.stderr)

def init_db():
    fresh = not os.path.exists(DB_PATH)
    with db_lock:
        con = _conn()
        try:
            con.executescript(_SCHEMA)
            _run_versioned_migrations(con)
            # A-01 合同编号唯一约束（数据库层兜底）：先扫描存量重复编号，存在则终止并输出清单
            dup_nos = con.execute('SELECT "编号", COUNT(*) c FROM contracts GROUP BY "编号" HAVING c>1').fetchall()
            if dup_nos:
                lst = "、".join(f"{r[0]}（{r[1]}条）" for r in dup_nos)
                print(f"[ERROR] contracts 存在重复编号，无法建立唯一索引，请先处理：{lst}", file=sys.stderr)
                raise RuntimeError("contracts 存在重复编号，无法建立唯一索引：" + lst)
            con.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_contracts_no ON contracts("编号")')
            # A-02 departments 复合主键（单位名称, 科室名称）数据库层兜底：重建表，先校验存量无重复组合
            dept_pk = [r[5] for r in con.execute("PRAGMA table_info(departments)").fetchall()]
            if not all(dept_pk):
                dup_depts = con.execute('SELECT "单位名称","科室名称",COUNT(*) c FROM departments GROUP BY "单位名称","科室名称" HAVING c>1').fetchall()
                if dup_depts:
                    lst = "、".join(f"{r[0]}-{r[1]}（{r[2]}条）" for r in dup_depts)
                    print(f"[ERROR] departments 存在重复部门组合，无法建立复合主键，请先处理：{lst}", file=sys.stderr)
                    raise RuntimeError("departments 存在重复部门组合，无法建立复合主键：" + lst)
                old_ddl = con.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='departments'").fetchone()[0]
                old_rows = con.execute('SELECT "单位名称","科室名称" FROM departments').fetchall()
                backup_path = os.path.join(os.path.dirname(DB_PATH), "departments_ddl_backup.sql")
                with open(backup_path, "w", encoding="utf-8", newline="\n") as f:
                    f.write(f"-- departments 迁移前原表结构与数据备份：{datetime.datetime.now():%Y-%m-%d %H:%M:%S}\n")
                    f.write(old_ddl.rstrip().rstrip(";") + ";\n")
                    for r in old_rows:
                        f.write('INSERT INTO departments ("单位名称","科室名称") VALUES (\'{}\',\'{}\');\n'.format(
                            str(r[0] or "").replace("'", "''"), str(r[1] or "").replace("'", "''")))
                con.execute("BEGIN")
                try:
                    # 修复后全量校验：contracts/users/attachments 任一悬空引用即终止（禁止静默跳过）
                    con.execute('CREATE TABLE departments_new ("单位名称" TEXT NOT NULL, "科室名称" TEXT NOT NULL, PRIMARY KEY ("单位名称","科室名称"))')
                    con.execute('INSERT INTO departments_new ("单位名称","科室名称") SELECT "单位名称","科室名称" FROM departments')
                    con.execute("DROP TABLE departments")
                    con.execute("ALTER TABLE departments_new RENAME TO departments")
                    con.execute("COMMIT")
                except Exception:
                    con.execute("ROLLBACK")
                    raise

            # A-03 外键约束（数据库层兜底）：重建 contracts/users/attachments 三张表并声明外键；先修复缺科室后缀数据
            ccols3 = [r[1] for r in con.execute("PRAGMA table_info(contracts)").fetchall()]
            if "所属单位" not in ccols3:
                dept_pairs = set((r[0], r[1]) for r in con.execute('SELECT "单位名称","科室名称" FROM departments').fetchall())
                def _split_dept(v):
                    v = str(v or "").strip()
                    if "-" in v:
                        u, k = v.split("-", 1)
                        return (u.strip(), k.strip())
                    return None
                broken = []
                for r in con.execute('SELECT id,"编号","项目名称","经办人","所属部门" FROM contracts ORDER BY "编号"').fetchall():
                    sp = _split_dept(r[4])
                    if sp is None:
                        if str(r[4] or "").strip(): broken.append(r)
                    elif sp not in dept_pairs:
                        broken.append(r)
                _DEPT_RULES = [
                    (("公车","车辆","平台委托","平台服务","新能源"), "公车管理组"),
                    (("周转房","周转住房"), "周转住房专班"),
                    (("施工图","预算","结算","审核","设计","幕墙","桥梁","钢结构","给水","钢筋","建设工程","工程质量","项目检测","检测合同","施工"), "工程建设组"),
                    (("物业","食堂","电梯","直饮水","消防","空调","窗帘","消杀","保洁","环境治理","标识","标牌","车库","会议室","弱电","网络","监控","配电","排污","管道","维保","维修","维护","运维","灭火器","卷帘门","限高","照明","下水道","水质","监测","采购","物资"), "物业管理组"),
                    (("市民之家",), "市民之家专班"),
                ]
                def _suggest_dept(name):
                    for kws, dept in _DEPT_RULES:
                        if any(k in (name or "") for k in kws): return dept, kws
                    return "后勤管理组", ("兜底",)
                repair_rows = []
                if broken:
                    md = ["# A-03 部门数据修复清单（缺科室后缀 → 建议科室）", "",
                          "> 依据关键字规则自动生成，**待业务逐条复核确认**；如需调整，按编号修改“所属部门”后重新执行迁移（回滚见 backups 重建前备份）。",
                          "", "| 编号 | 项目名称 | 原所属部门 | 修复后所属部门 | 修复依据（命中关键字） |", "|---|---|---|---|---|"]
                    for r in broken:
                        dept, kws = _suggest_dept(r[2])
                        newval = (str(r[4] or "").strip()) + "-" + dept
                        repair_rows.append((r[0], newval))
                        md.append("| {} | {} | {} | {} | {} |".format(
                            str(r[1]).replace("|", "\\|"), str(r[2] or "").replace("|", "\\|").replace("\n", " "),
                            str(r[4] or "").replace("|", "\\|"), newval.replace("|", "\\|"), "、".join(kws)))
                    md_path = os.path.join(BASE_DIR, "docs", "A-03-部门数据修复清单.md")
                    os.makedirs(os.path.dirname(md_path), exist_ok=True)
                    with open(md_path, "w", encoding="utf-8", newline="\n") as f:
                        f.write("\n".join(md) + "\n")
                    print(f"[INFO] A-03 缺科室后缀部门数据 {len(broken)} 条，已按建议修复，清单：{md_path}", file=sys.stderr)
                # 重建前整库备份
                os.makedirs(BACKUP_DIR, exist_ok=True)
                ts1 = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                bak_before = os.path.join(BACKUP_DIR, f"合同台账_A03_重建前_{ts1}.db")
                if os.path.exists(bak_before): os.remove(bak_before)
                con.execute("VACUUM INTO '" + bak_before.replace("'", "''") + "'")
                # 事务内：先修复缺科室后缀数据并全量校验（任一悬空引用即终止），再原子重建三张表
                con.execute("BEGIN")
                try:
                    for rid, newval in repair_rows:
                        con.execute('UPDATE contracts SET "所属部门"=? WHERE id=?', (newval, rid))
                    bad_dept = []
                    for r in con.execute('SELECT id,"编号","所属部门" FROM contracts').fetchall():
                        sp = _split_dept(r[2])
                        if sp is None:
                            if str(r[2] or "").strip(): bad_dept.append(r[1])
                        elif sp not in dept_pairs:
                            bad_dept.append(r[1])
                    if bad_dept:
                        raise RuntimeError("contracts 存在无法匹配部门的所属部门，请先修复：" + "、".join(str(x) for x in bad_dept[:50]))
                    bad_user = []
                    for r in con.execute('SELECT "用户名","所属部门" FROM users').fetchall():
                        sp = _split_dept(r[1])
                        if sp is None or sp not in dept_pairs:
                            bad_user.append(str(r[0]))
                    if bad_user:
                        raise RuntimeError("users 存在无法匹配部门的所属部门，请先修复：" + "、".join(bad_user))
                    cno_set = set(x[0] for x in con.execute('SELECT "编号" FROM contracts').fetchall())
                    bad_att = []
                    for r in con.execute('SELECT "合同编号" FROM attachments').fetchall():
                        if str(r[0] or "").strip() and str(r[0]) not in cno_set:
                            bad_att.append(str(r[0]))
                    if bad_att:
                        raise RuntimeError("attachments 存在不存在的合同编号，请先修复：" + "、".join(bad_att))
                    def _mk_new_ddl(con, tbl, new_name, extra_cols, fk_clause):
                        info = con.execute("PRAGMA table_info(" + tbl + ")").fetchall()
                        segs = []
                        for cid, cname, ctype, notnull, dflt, pk in info:
                            seg = '"%s" %s' % (cname, ctype or "TEXT")
                            if notnull: seg += " NOT NULL"
                            if dflt is not None: seg += " DEFAULT " + str(dflt)
                            if pk == 1:
                                if cname == "id" and (ctype or "").upper() == "INTEGER":
                                    seg = '"id" INTEGER PRIMARY KEY AUTOINCREMENT'
                                else:
                                    seg += " PRIMARY KEY"
                            segs.append(seg)
                        if extra_cols: segs.append(extra_cols)
                        if fk_clause: segs.append(fk_clause)
                        return "CREATE TABLE " + new_name + " (\n  " + ",\n  ".join(segs) + "\n)"
                    # contracts：追加 所属单位/所属科室 + 复合外键 → departments
                    old_ccols = [r[1] for r in con.execute("PRAGMA table_info(contracts)").fetchall()]
                    c_ddl = _mk_new_ddl(con, "contracts", "contracts_new",
                                        '"所属单位" TEXT, "所属科室" TEXT',
                                        'FOREIGN KEY ("所属单位","所属科室") REFERENCES departments("单位名称","科室名称") ON DELETE RESTRICT ON UPDATE CASCADE')
                    con.execute(c_ddl)
                    csel = ",".join('"%s"' % c for c in old_ccols)
                    con.execute("INSERT INTO contracts_new (" + csel + ") SELECT " + csel + " FROM contracts")
                    con.execute('''UPDATE contracts_new SET
  "所属单位" = CASE WHEN instr("所属部门",'-')>0 THEN substr("所属部门",1,instr("所属部门",'-')-1) ELSE '' END,
  "所属科室" = CASE WHEN instr("所属部门",'-')>0 THEN substr("所属部门",instr("所属部门",'-')+1) ELSE '' END
  WHERE "所属部门" IS NOT NULL AND TRIM("所属部门") <> '' ''')
                    # DROP/重命名父表前，先移除引用 contracts 的触发器（否则 RENAME 时触发解析报错），
                    # 事务提交后由下方 A-03 触发器块 CREATE TRIGGER IF NOT EXISTS 自动重建
                    con.execute("DROP TRIGGER IF EXISTS trg_contracts_dept_sync_i")
                    con.execute("DROP TRIGGER IF EXISTS trg_contracts_dept_sync_u")
                    con.execute("DROP TRIGGER IF EXISTS trg_departments_rename_sync")
                    con.execute("DROP TABLE contracts")
                    con.execute("ALTER TABLE contracts_new RENAME TO contracts")
                    con.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_contracts_no ON contracts("编号")')
                    # users：追加 所属单位/所属科室 + 复合外键 → departments
                    old_ucols = [r[1] for r in con.execute("PRAGMA table_info(users)").fetchall()]
                    u_ddl = _mk_new_ddl(con, "users", "users_new",
                                        '"所属单位" TEXT, "所属科室" TEXT',
                                        'FOREIGN KEY ("所属单位","所属科室") REFERENCES departments("单位名称","科室名称") ON DELETE RESTRICT ON UPDATE CASCADE')
                    con.execute(u_ddl)
                    usel = ",".join('"%s"' % c for c in old_ucols)
                    con.execute("INSERT INTO users_new (" + usel + ") SELECT " + usel + " FROM users")
                    con.execute('''UPDATE users_new SET
  "所属单位" = CASE WHEN instr("所属部门",'-')>0 THEN substr("所属部门",1,instr("所属部门",'-')-1) ELSE '' END,
  "所属科室" = CASE WHEN instr("所属部门",'-')>0 THEN substr("所属部门",instr("所属部门",'-')+1) ELSE '' END
  WHERE "所属部门" IS NOT NULL AND TRIM("所属部门") <> '' ''')
                    con.execute("DROP TABLE users")
                    con.execute("ALTER TABLE users_new RENAME TO users")
                    # attachments：追加外键 → contracts(编号)
                    old_acols = [r[1] for r in con.execute("PRAGMA table_info(attachments)").fetchall()]
                    a_ddl = _mk_new_ddl(con, "attachments", "attachments_new",
                                        None,
                                        'FOREIGN KEY ("合同编号") REFERENCES contracts("编号") ON DELETE RESTRICT ON UPDATE RESTRICT')
                    con.execute(a_ddl)
                    asel = ",".join('"%s"' % c for c in old_acols)
                    con.execute("INSERT INTO attachments_new (" + asel + ") SELECT " + asel + " FROM attachments")
                    con.execute("DROP TABLE attachments")
                    con.execute("ALTER TABLE attachments_new RENAME TO attachments")
                    con.execute("COMMIT")
                except Exception:
                    con.execute("ROLLBACK")
                    raise
                finally:
                    con.execute("PRAGMA foreign_keys=ON")
                # 重建后整库备份
                ts2 = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                bak_after = os.path.join(BACKUP_DIR, f"合同台账_A03_重建后_{ts2}.db")
                if os.path.exists(bak_after): os.remove(bak_after)
                con.execute("VACUUM INTO '" + bak_after.replace("'", "''") + "'")
            # A-05 CHECK 与 NOT NULL 约束（金额非负/日期格式/必填字段）：重建 contracts；枚举 CHECK 延后至 A-08 清洗后启用
            cinfo5 = con.execute("PRAGMA table_info(contracts)").fetchall()
            cnames5 = [r[1] for r in cinfo5]
            cnot5 = [r[3] for r in cinfo5]
            if "编号" in cnames5 and not cnot5[cnames5.index("编号")]:
                req5 = ["编号","项目分类","项目名称","采购方式","合同金额（万元）","资金来源","合同主要内容","经办人","所属部门"]
                amt5 = ["合同金额（万元）","结算金额","已支付","未支付","决策的招标控制价"]
                date5 = ["采购时间","开工时间","完工时间","竣工验收时间","下次招采时间"]
                import re as _re5
                def _norm_date5(v):
                    s = str(v or "").strip()
                    if not s: return ""
                    m = _re5.search(r"(20\d{2})[.\-\s年]+(\d{1,2})[.\-\s月]+(\d{1,2})", s)
                    if m: return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
                    return ""
                def _extract_num5(v):
                    # 金额文本解析：纯数字直接转换；否则提取首个数字作为“建议值”参考；无法解析返回 None
                    if v is None: return None
                    s = str(v).strip()
                    if not s: return None
                    try:
                        return float(s)
                    except ValueError:
                        m = _re5.search(r"-?\d+(?:\.\d+)?", s)
                        return float(m.group(0)) if m else None
                def _sib_handler5(no):
                    base = str(no).split("-")[0]
                    for pfx in (base, base[:5]):
                        row = con.execute('SELECT "经办人" FROM contracts WHERE "编号" LIKE ? AND "经办人" IS NOT NULL AND TRIM(CAST("经办人" AS TEXT))<>"" ORDER BY "编号" LIMIT 1', (pfx + "%",)).fetchone()
                        if row: return str(row[0])
                    return ""
                repair5 = []  # (id, 编号, 列, 原值, 新值, 依据)：本次自动应用
                for r in con.execute('SELECT id,"编号","项目分类","项目名称","采购方式","合同金额（万元）","资金来源","合同主要内容","经办人" FROM contracts ORDER BY id').fetchall():
                    rid, no = r[0], str(r[1])
                    def g5(v): return "" if v is None else str(v).strip()
                    if not g5(r[2]): repair5.append((rid, no, "项目分类", "", "待补充", "必填字段为空，回填占位值"))
                    if not g5(r[3]): repair5.append((rid, no, "项目名称", "", "待补充", "必填字段为空，回填占位值"))
                    if not g5(r[4]): repair5.append((rid, no, "采购方式", "", "自行采购", "必填字段为空，按默认枚举回填"))
                    if r[5] is None or g5(r[5])=="": repair5.append((rid, no, "合同金额（万元）", "", 0, "必填字段为空，按 0 回填"))
                    if not g5(r[6]): repair5.append((rid, no, "资金来源", "", "其他", "必填字段为空，按兜底枚举“其他”回填"))
                    if not g5(r[7]): repair5.append((rid, no, "合同主要内容", "", "待补充", "必填字段为空，回填占位值"))
                    if not g5(r[8]):
                        sh = _sib_handler5(no)
                        repair5.append((rid, no, "经办人", "", sh if sh else "待补充", "必填字段为空，优先取同前缀兄弟合同经办人，否则回填占位值"))
                for r in con.execute('SELECT id,"编号","采购时间","开工时间","完工时间","竣工验收时间","下次招采时间" FROM contracts ORDER BY id').fetchall():
                    rid, no = r[0], str(r[1])
                    for ci, col in enumerate(date5):
                        v = r[2 + ci]
                        if v is None or str(v).strip() == "": continue
                        sv = str(v).strip()
                        if _re5.match(r"^\d{4}-\d{2}-\d{2}$", sv): continue
                        nv = _norm_date5(sv)
                        repair5.append((rid, no, col, sv, nv, "日期格式非 YYYY-MM-DD，提取日期或清空"))
                # 金额列非数字文本：按金额 CHECK（CAST(字段 AS REAL)>=0）语义，文本 CAST 后为 0、满足约束，
                # 故迁移不阻塞、不自动改写（避免将文号“潭财办发[2014]15号”、年份“2025年”、单位说明等误判为金额），仅输出清单供业务复核。
                amt_warn5 = []  # (id, 编号, 列, 原值, 建议值)
                for r in con.execute('SELECT id,"编号","合同金额（万元）","结算金额","已支付","未支付","决策的招标控制价" FROM contracts ORDER BY id').fetchall():
                    rid, no = r[0], str(r[1])
                    for ci, col in enumerate(amt5):
                        v = r[2 + ci]
                        if v is None or str(v).strip() == "": continue
                        try:
                            float(str(v).strip())
                            continue
                        except ValueError:
                            pass
                        sug = _extract_num5(v)
                        amt_warn5.append((rid, no, col, str(v).strip(), sug if sug is not None else "无法解析"))
                if repair5 or amt_warn5:
                    md5 = ["# A-05 数据修复清单（CHECK/NOT NULL 前置：存量违规逐条回填）", "",
                           "> 依据明确规则自动修复，**待业务复核确认**；回滚见 backups A-05 重建前备份。",
                           "> 枚举 CHECK（项目分类/采购方式/是否结算/资金来源/是否归档）延后至 A-08 清洗后启用，本次未启用。",
                           "", "## 一、自动回填（本次迁移已应用）", "",
                           "| 编号 | 字段 | 原值 | 修复值 | 修复依据 |", "|---|---|---|---|---|"]
                    for rid, no, col, old, new, why in repair5:
                        md5.append("| {} | {} | {} | {} | {} |".format(str(no).replace("|", "\\|").replace("\n", " "), str(col).replace("|", "\\|").replace("\n", " "), str(old).replace("|", "\\|").replace("\n", " "), str(new).replace("|", "\\|").replace("\n", " "), str(why).replace("|", "\\|").replace("\n", " ")))
                    md5 += ["", "## 二、金额列非数字文本（建议人工复核，本次未自动修改）", "",
                            "> 按 A-05 金额 CHECK（CAST(字段 AS REAL)>=0）语义，以下文本 CAST 后为 0，满足约束，故迁移不阻塞、不改写。",
                            "> 建议业务逐条核实后手动修正；“建议值”仅为正则提取的首个数字参考，不构成金额换算。",
                            "", "| id | 编号 | 字段 | 原值 | 建议值 |", "|---|---|---|---|---|"]
                    for rid, no, col, old, sug in amt_warn5:
                        md5.append("| {} | {} | {} | {} | {} |".format(rid, str(no).replace("|", "\\|").replace("\n", " "), str(col).replace("|", "\\|").replace("\n", " "), str(old).replace("|", "\\|").replace("\n", " "), str(sug).replace("|", "\\|").replace("\n", " ")))
                    md5_path = os.path.join(BASE_DIR, "docs", "A-05-数据修复清单.md")
                    os.makedirs(os.path.dirname(md5_path), exist_ok=True)
                    with open(md5_path, "w", encoding="utf-8", newline="\n") as f:
                        f.write("\n".join(md5) + "\n")
                    print(f"[INFO] A-05 自动回填 {len(repair5)} 条；金额文本待复核 {len(amt_warn5)} 条，清单：{md5_path}", file=sys.stderr)
                # 重建前整库备份
                os.makedirs(BACKUP_DIR, exist_ok=True)
                ts5 = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                bak5_before = os.path.join(BACKUP_DIR, f"合同台账_A05_重建前_{ts5}.db")
                if os.path.exists(bak5_before): os.remove(bak5_before)
                con.execute("VACUUM INTO '" + bak5_before.replace("'", "''") + "'")
                # 事务内：回填违规数据 → 复验零违规 → 重建 contracts（含 CHECK/NOT NULL）
                # 重建需 DROP 被 attachments 外键引用的 contracts 父表；外键开关仅对当前连接生效，
                # 临时关闭（PRAGMA 须在事务外执行），事务结束后恢复；API 连接(_conn)仍强制外键。
                con.execute("PRAGMA foreign_keys=OFF")
                con.execute("BEGIN")
                try:
                    for rid, no, col, old, new, why in repair5:
                        con.execute('UPDATE contracts SET "' + col + '"=? WHERE id=?', (new, rid))
                    bad5 = []
                    for r in con.execute('SELECT id,"编号","项目分类","项目名称","采购方式","合同金额（万元）","资金来源","合同主要内容","经办人","所属部门" FROM contracts').fetchall():
                        if any((r[i] is None or str(r[i]).strip()=="") for i in range(2, 10)): bad5.append((str(r[1]), "必填为空"))
                    for r in con.execute('SELECT id,"编号","合同金额（万元）","结算金额","已支付","未支付","决策的招标控制价" FROM contracts').fetchall():
                        for i, c in enumerate(amt5):
                            v = r[2 + i]
                            if v is None or str(v).strip() == "": continue
                            nf = _extract_num5(v)
                            # 非数字文本按 CHECK 语义（CAST 为 0）通过，仅负值视为违规
                            if nf is not None and nf < 0: bad5.append((str(r[1]), c + "为负"))
                    for r in con.execute('SELECT id,"编号","采购时间","开工时间","完工时间","竣工验收时间","下次招采时间" FROM contracts').fetchall():
                        for i, c in enumerate(date5):
                            sv = str(r[2 + i] or "").strip()
                            if sv and not _re5.match(r"^\d{4}-\d{2}-\d{2}$", sv): bad5.append((str(r[1]), c + "日期格式"))
                    if bad5:
                        raise RuntimeError("A-05 存量仍有违规，无法启用约束：" + "、".join(str(x) for x in bad5[:50]))
                    info5 = con.execute("PRAGMA table_info(contracts)").fetchall()
                    segs5 = []
                    for cid, cname, ctype, notnull, dflt, pk in info5:
                        seg = '"%s" %s' % (cname, ctype or "TEXT")
                        if cname in req5: seg += " NOT NULL"
                        if cname in amt5: seg += ' CHECK ("%s" IS NULL OR CAST("%s" AS REAL) >= 0)' % (cname, cname)
                        if cname in date5: seg += " CHECK (\"%s\" IS NULL OR \"%s\"='' OR \"%s\" GLOB '[0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]')" % (cname, cname, cname)
                        if dflt is not None: seg += " DEFAULT " + str(dflt)
                        if pk == 1:
                            if cname == "id" and (ctype or "").upper() == "INTEGER": seg = '"id" INTEGER PRIMARY KEY AUTOINCREMENT'
                            else: seg += " PRIMARY KEY"
                        segs5.append(seg)
                    segs5.append('FOREIGN KEY ("所属单位","所属科室") REFERENCES departments("单位名称","科室名称") ON DELETE RESTRICT ON UPDATE CASCADE')
                    con.execute("CREATE TABLE contracts_new (\n  " + ",\n  ".join(segs5) + "\n)")
                    old5 = [r[1] for r in con.execute("PRAGMA table_info(contracts)").fetchall()]
                    sel5 = ",".join('"%s"' % c for c in old5)
                    con.execute("INSERT INTO contracts_new (" + sel5 + ") SELECT " + sel5 + " FROM contracts")
                    # DROP/重命名父表前，先移除引用 contracts 的触发器（否则 RENAME 时触发解析报错），
                    # 事务提交后由下方 A-03 触发器块 CREATE TRIGGER IF NOT EXISTS 自动重建
                    con.execute("DROP TRIGGER IF EXISTS trg_contracts_dept_sync_i")
                    con.execute("DROP TRIGGER IF EXISTS trg_contracts_dept_sync_u")
                    con.execute("DROP TRIGGER IF EXISTS trg_departments_rename_sync")
                    con.execute("DROP TABLE contracts")
                    con.execute("ALTER TABLE contracts_new RENAME TO contracts")
                    con.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_contracts_no ON contracts("编号")')
                    con.execute("COMMIT")
                except Exception:
                    con.execute("ROLLBACK")
                    raise
                finally:
                    con.execute("PRAGMA foreign_keys=ON")
                # 重建后整库备份
                ts52 = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                bak5_after = os.path.join(BACKUP_DIR, f"合同台账_A05_重建后_{ts52}.db")
                if os.path.exists(bak5_after): os.remove(bak5_after)
                con.execute("VACUUM INTO '" + bak5_after.replace("'", "''") + "'")

            # A-06 字段类型规范化：attachments.文件大小 TEXT→INTEGER；日期列统一 YYYY-MM-DD；
            # 金额列声明类型已为 REAL，存量文本值须人工确认（迁移不猜测转换），输出清单供业务复核
            import re as _re6
            # ── 1) attachments.文件大小 → INTEGER：重建子表（保留主键与外键），存量字节数字串 CAST ──
            a_info6 = con.execute("PRAGMA table_info(attachments)").fetchall()
            a_types6 = {r[1]: r[2] for r in a_info6}
            if str(a_types6.get("文件大小", "") or "").upper() != "INTEGER":
                os.makedirs(BACKUP_DIR, exist_ok=True)
                ts6 = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                bak6 = os.path.join(BACKUP_DIR, f"合同台账_A06_附件重建前_{ts6}.db")
                if os.path.exists(bak6): os.remove(bak6)
                con.execute("VACUUM INTO '" + bak6.replace("'", "''") + "'")
                con.execute("BEGIN")
                try:
                    segs6 = []
                    for cid6, cn6, ct6, nn6, df6, pk6 in a_info6:
                        seg = '"%s" %s' % (cn6, "INTEGER" if cn6 == "文件大小" else (ct6 or "TEXT"))
                        if df6 is not None: seg += " DEFAULT " + str(df6)
                        if pk6 == 1: seg += " PRIMARY KEY"
                        segs6.append(seg)
                    segs6.append('FOREIGN KEY ("合同编号") REFERENCES contracts("编号") ON DELETE RESTRICT ON UPDATE RESTRICT')
                    con.execute("CREATE TABLE attachments_new (\n  " + ",\n  ".join(segs6) + "\n)")
                    aold6 = [r[1] for r in a_info6]
                    asel6 = ",".join('"%s"' % c for c in aold6)
                    acast6 = ",".join('CAST("文件大小" AS INTEGER)' if c == "文件大小" else '"%s"' % c for c in aold6)
                    con.execute("INSERT INTO attachments_new (" + asel6 + ") SELECT " + acast6 + " FROM attachments")
                    con.execute("DROP TABLE attachments")
                    con.execute("ALTER TABLE attachments_new RENAME TO attachments")
                    con.execute("COMMIT")
                except Exception:
                    con.execute("ROLLBACK")
                    raise
                ts62 = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                bak6b = os.path.join(BACKUP_DIR, f"合同台账_A06_附件重建后_{ts62}.db")
                if os.path.exists(bak6b): os.remove(bak6b)
                con.execute("VACUUM INTO '" + bak6b.replace("'", "''") + "'")
                print("[INFO] A-06 attachments.文件大小 已重建为 INTEGER", file=sys.stderr)
            # ── 2) 日期列统一 YYYY-MM-DD（含归档时间）；无法解析的仅列清单，禁止静默置空 ──
            date6 = ["采购时间","开工时间","完工时间","竣工验收时间","下次招采时间","归档时间"]
            def _norm_date6(v):
                s = str(v or "").strip()
                if not s: return None
                m = _re6.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})([ T].*)?$", s)
                if m: return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
                m = _re6.search(r"(20\d{2})[.\-\s年]+(\d{1,2})[.\-\s月]+(\d{1,2})", s)
                if m: return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
                return None
            def _extract_num6(v):
                if v is None: return None
                s = str(v).strip()
                if not s: return None
                try:
                    return float(s)
                except ValueError:
                    m = _re6.search(r"-?\d+(?:\.\d+)?", s)
                    return float(m.group(0)) if m else None
            date_fix6 = []   # (编号, 列, 原值, 新值)：自动规范化
            date_bad6 = []   # (编号, 列, 原值)：无法解析，待业务确认
            for r in con.execute('SELECT id,"编号","采购时间","开工时间","完工时间","竣工验收时间","下次招采时间","归档时间" FROM contracts').fetchall():
                rid, no = r[0], str(r[1])
                for ci, col in enumerate(date6):
                    v = r[2 + ci]
                    if v is None or str(v).strip() == "": continue
                    sv = str(v).strip()
                    if _re6.match(r"^\d{4}-\d{2}-\d{2}$", sv): continue
                    nv = _norm_date6(sv)
                    if nv is None: date_bad6.append((no, col, sv))
                    else: date_fix6.append((no, col, sv, nv))
            if date_fix6:
                con.execute("BEGIN")
                try:
                    for no, col, old, nv in date_fix6:
                        con.execute('UPDATE contracts SET "' + col + '"=? WHERE "编号"=?', (nv, no))
                    con.execute("COMMIT")
                except Exception:
                    con.execute("ROLLBACK")
                    raise
            md6 = ["# A-06 日期规范化清单（统一 YYYY-MM-DD）", "",
                   "> 自动规范化已应用；无法解析的日期仅列出，**待业务确认后回填**，未静默置空。",
                   "", "## 一、自动规范化（本次迁移已应用）", "",
                   "| 编号 | 字段 | 原值 | 新值 |", "|---|---|---|---|"]
            for no, col, old, nv in date_fix6:
                md6.append("| {} | {} | {} | {} |".format(str(no).replace("|", "\\|").replace("\n", " "), col, old.replace("|", "\\|").replace("\n", " "), nv))
            md6 += ["", "## 二、无法解析（待业务确认回填，未修改）", "",
                    "| 编号 | 字段 | 原值 |", "|---|---|---|"]
            for no, col, old in date_bad6:
                md6.append("| {} | {} | {} |".format(str(no).replace("|", "\\|").replace("\n", " "), col, old.replace("|", "\\|").replace("\n", " ")))
            md6_path = os.path.join(BASE_DIR, "docs", "A-06-日期规范化清单.md")
            os.makedirs(os.path.dirname(md6_path), exist_ok=True)
            with open(md6_path, "w", encoding="utf-8", newline="\n") as f:
                f.write("\n".join(md6) + "\n")
            print(f"[INFO] A-06 日期自动规范化 {len(date_fix6)} 条；无法解析待确认 {len(date_bad6)} 条，清单：{md6_path}", file=sys.stderr)
            # ── 3) 金额列文本：声明类型已为 REAL；存量文本值迁移不猜测转换，输出清单待业务确认 ──
            amt6 = ["合同金额（万元）","结算金额","已支付","未支付","决策的招标控制价"]
            amt_warn6 = []  # (id, 编号, 列, 原值, 建议值)
            for r in con.execute('SELECT id,"编号","合同金额（万元）","结算金额","已支付","未支付","决策的招标控制价" FROM contracts ORDER BY id').fetchall():
                rid, no = r[0], str(r[1])
                for ci, col in enumerate(amt6):
                    v = r[2 + ci]
                    if v is None or str(v).strip() == "": continue
                    try:
                        float(str(v).strip())
                        continue
                    except ValueError:
                        pass
                    sug = _extract_num6(v)
                    amt_warn6.append((rid, no, col, str(v).strip(), sug if sug is not None else "无法解析"))
            if amt_warn6:
                md6a = ["# A-06 金额类型转换确认清单（存量文本值，待业务确认）", "",
                        "> 实况：金额列声明类型已为 REAL（合同金额/结算金额/已支付/未支付/决策的招标控制价），残留问题仅是**存量文本值**。",
                        "> 迁移脚本**不做猜测转换**（避免把文号/年份/单位说明误判为金额）；请业务逐条核对正确金额后，",
                        "> 通过系统“修改合同”表单或 SQL 更新为数字（写入 REAL 列时自动转为数值）。",
                        "> “建议值”仅为正则提取的首个数字参考，不构成金额换算。",
                        "", "| id | 编号 | 字段 | 原值 | 建议值 | 状态 |", "|---|---|---|---|---|---|"]
                for rid, no, col, old, sug in amt_warn6:
                    md6a.append("| {} | {} | {} | {} | {} | 待确认 |".format(rid, str(no).replace("|", "\\|").replace("\n", " "), col, old.replace("|", "\\|").replace("\n", " "), str(sug).replace("|", "\\|").replace("\n", " ")))
                md6a_path = os.path.join(BASE_DIR, "docs", "A-06-金额类型转换确认清单.md")
                os.makedirs(os.path.dirname(md6a_path), exist_ok=True)
                with open(md6a_path, "w", encoding="utf-8", newline="\n") as f:
                    f.write("\n".join(md6a) + "\n")
                print(f"[INFO] A-06 金额列文本值 {len(amt_warn6)} 条待业务确认，清单：{md6a_path}", file=sys.stderr)
            else:
                print("[INFO] A-06 金额列无文本残留，无需转换", file=sys.stderr)
            # A-03 同步触发器：所属部门(字符串) ↔ 所属单位/所属科室(外键列)，保证 API 写入一致；部门改名级联同步
            con.execute('''CREATE TRIGGER IF NOT EXISTS trg_contracts_dept_sync_i
AFTER INSERT ON contracts
FOR EACH ROW WHEN (NEW."所属部门" IS NOT NULL AND TRIM(NEW."所属部门") <> '')
BEGIN
  UPDATE contracts SET
    "所属单位" = CASE WHEN instr(NEW."所属部门",'-')>0 THEN substr(NEW."所属部门",1,instr(NEW."所属部门",'-')-1) ELSE '' END,
    "所属科室" = CASE WHEN instr(NEW."所属部门",'-')>0 THEN substr(NEW."所属部门",instr(NEW."所属部门",'-')+1) ELSE '' END
  WHERE id = NEW.id;
END''')
            con.execute('''CREATE TRIGGER IF NOT EXISTS trg_contracts_dept_sync_u
AFTER UPDATE OF "所属部门" ON contracts
FOR EACH ROW WHEN (NEW."所属部门" IS NOT NULL AND TRIM(NEW."所属部门") <> '')
BEGIN
  UPDATE contracts SET
    "所属单位" = CASE WHEN instr(NEW."所属部门",'-')>0 THEN substr(NEW."所属部门",1,instr(NEW."所属部门",'-')-1) ELSE '' END,
    "所属科室" = CASE WHEN instr(NEW."所属部门",'-')>0 THEN substr(NEW."所属部门",instr(NEW."所属部门",'-')+1) ELSE '' END
  WHERE id = NEW.id;
END''')
            con.execute('''CREATE TRIGGER IF NOT EXISTS trg_users_dept_sync_i
AFTER INSERT ON users
FOR EACH ROW WHEN (NEW."所属部门" IS NOT NULL AND TRIM(NEW."所属部门") <> '')
BEGIN
  UPDATE users SET
    "所属单位" = CASE WHEN instr(NEW."所属部门",'-')>0 THEN substr(NEW."所属部门",1,instr(NEW."所属部门",'-')-1) ELSE '' END,
    "所属科室" = CASE WHEN instr(NEW."所属部门",'-')>0 THEN substr(NEW."所属部门",instr(NEW."所属部门",'-')+1) ELSE '' END
  WHERE "用户名" = NEW."用户名";
END''')
            con.execute('''CREATE TRIGGER IF NOT EXISTS trg_users_dept_sync_u
AFTER UPDATE OF "所属部门" ON users
FOR EACH ROW WHEN (NEW."所属部门" IS NOT NULL AND TRIM(NEW."所属部门") <> '')
BEGIN
  UPDATE users SET
    "所属单位" = CASE WHEN instr(NEW."所属部门",'-')>0 THEN substr(NEW."所属部门",1,instr(NEW."所属部门",'-')-1) ELSE '' END,
    "所属科室" = CASE WHEN instr(NEW."所属部门",'-')>0 THEN substr(NEW."所属部门",instr(NEW."所属部门",'-')+1) ELSE '' END
  WHERE "用户名" = NEW."用户名";
END''')
            con.execute('''CREATE TRIGGER IF NOT EXISTS trg_departments_rename_sync
AFTER UPDATE OF "单位名称","科室名称" ON departments
FOR EACH ROW
BEGIN
  UPDATE contracts SET "所属部门" = NEW."单位名称" || '-' || NEW."科室名称"
    WHERE "所属部门" = OLD."单位名称" || '-' || OLD."科室名称";
  UPDATE users SET "所属部门" = NEW."单位名称" || '-' || NEW."科室名称"
    WHERE "所属部门" = OLD."单位名称" || '-' || OLD."科室名称";
END''')

            # A-04 contracts 常用查询列索引（幂等；编号唯一索引 A-01 已建，直接复用不重复创建）
            # 经办人为多值列（顿号/斜杠分隔多人），索引价值有限，作为过渡索引；档位 B 拆分多值列后将废弃本索引
            con.execute('CREATE INDEX IF NOT EXISTS idx_contracts_department ON contracts("所属部门")')
            con.execute('CREATE INDEX IF NOT EXISTS idx_contracts_procurement_date ON contracts("采购时间")')
            con.execute('CREATE INDEX IF NOT EXISTS idx_contracts_next_bid_date ON contracts("下次招采时间")')
            con.execute('CREATE INDEX IF NOT EXISTS idx_contracts_archived ON contracts("是否归档")')
            con.execute('CREATE INDEX IF NOT EXISTS idx_contracts_handler ON contracts("经办人")')

            con.execute('INSERT OR IGNORE INTO settings ("配置项","配置值") VALUES (?,?)', ("log_keep_days", "90"))


            con.execute('INSERT OR IGNORE INTO settings ("配置项","配置值") VALUES (?,?)', ("default_attach_max_mb", "20"))
            con.commit()
        finally:
            con.close()
    if fresh:
        if os.path.exists(EXCEL_PATH):
            _migrate_from_excel()
        _seed_defaults()

def _norm(v):
    if v is None: return ""
    if isinstance(v, datetime.datetime): return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date): return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)): return v
    return str(v).strip()

def _insert_dict(con, table, colnames, rowdict):
    cols = list(colnames)
    sql = "INSERT INTO " + table + " (" + ",".join('"'+c+'"' for c in cols) + ") VALUES (" + ",".join("?"*len(cols)) + ")"
    con.execute(sql, [rowdict.get(c, "") for c in cols])

def _migrate_from_excel():
    if not os.path.exists(EXCEL_PATH): return
    print("[INFO] 从 Excel 迁移数据到 SQLite...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    with db_lock:
        con = _conn()
        try:
            ws = wb["合同台账"]
            rows = list(ws.iter_rows(values_only=True))
            if rows and len(rows) > 1:
                hdr = [str(x).strip() if x else "" for x in rows[0]]
                for ri in range(1, len(rows)):
                    vals = {cn: "" for cn in CONTRACT_COLUMNS}
                    src = rows[ri]
                    for ci, cn in enumerate(hdr):
                        if cn == "\u7aef\u5de5\u9a8c\u6536\u65f6\u95f4": cn = "\u7ae3\u5de5\u9a8c\u6536\u65f6\u95f4"
                        if cn in ("经办人/分管领导","甲方现场负责人"):
                            cn = {"经办人/分管领导":"经办人","甲方现场负责人":"现场负责人"}[cn]
                        if cn in vals:
                            vals[cn] = _norm(src[ci]) if ci < len(src) else ""
                    cols = ["id"] + list(CONTRACT_COLUMNS)
                    sql = "INSERT INTO contracts (" + ",".join('"'+c+'"' for c in cols) + ") VALUES (" + ",".join("?"*len(cols)) + ")"
                    con.execute(sql, [ri+1] + [vals[cn] for cn in CONTRACT_COLUMNS])
            ws = wb["用户表"]
            rows = list(ws.iter_rows(values_only=True))
            if rows and len(rows) > 1:
                hdr = [str(x).strip() if x else "" for x in rows[0]]
                umap = {"用户名":"用户名","密码哈希":"密码","密码":"密码","姓名":"姓名","所属部门":"所属部门","角色":"角色"}
                for r in rows[1:]:
                    d = {}
                    for ci, hh in enumerate(hdr):
                        if hh in umap:
                            d[umap[hh]] = _norm(r[ci]) if ci < len(r) else ""
                    if not d.get("用户名"): continue
                    _insert_dict(con, "users", ["用户名","密码","姓名","所属部门","角色"], d)
            ws = wb["部门列表"]
            rows = list(ws.iter_rows(values_only=True))
            if rows and len(rows) > 1:
                hdr = [str(x).strip() if x else "" for x in rows[0]]
                for r in rows[1:]:
                    d = {hh: (_norm(r[i]) if i < len(r) else "") for i, hh in enumerate(hdr) if hh in ("单位名称","科室名称")}
                    if not d.get("单位名称") and not d.get("科室名称"): continue
                    _insert_dict(con, "departments", ["单位名称","科室名称"], d)
            ws = wb["系统配置"]
            rows = list(ws.iter_rows(values_only=True))
            if rows and len(rows) > 1:
                hdr = [str(x).strip() if x else "" for x in rows[0]]
                kcol = "配置项" if "配置项" in hdr else ("键" if "键" in hdr else None)
                vcol = "配置值" if "配置值" in hdr else ("值" if "值" in hdr else None)
                if kcol and vcol:
                    ki = hdr.index(kcol); vi = hdr.index(vcol)
                    for r in rows[1:]:
                        k = _norm(r[ki]) if ki < len(r) else ""
                        v = _norm(r[vi]) if vi < len(r) else ""
                        if k:
                            con.execute('INSERT INTO settings ("配置项","配置值") VALUES (?,?)', (k, v))
            ws = wb[ATTACH_SHEET]
            rows = list(ws.iter_rows(values_only=True))
            if rows and len(rows) > 1:
                hdr = [str(x).strip() if x else "" for x in rows[0]]
                for r in rows[1:]:
                    d = {hh: (_norm(r[i]) if i < len(r) else "") for i, hh in enumerate(hdr) if hh in ATTACH_HEADERS}
                    if not d.get("附件ID"): continue
                    _insert_dict(con, "attachments", ATTACH_HEADERS, d)
            con.commit()
        except Exception:
            con.rollback(); raise
        finally:
            con.close()
    print("[INFO] 迁移完成。")

def _seed_defaults():
    with db_lock:
        con = _conn()
        try:
            # 注意：必须先插入部门再插入用户——用户表含外键(所属单位,所属科室)→departments，
            # 且 trg_users_dept_sync_i 触发器会回填外键列，先存在部门表中才能通过外键校验。
            if con.execute('SELECT COUNT(*) FROM departments').fetchone()[0] == 0:
                con.execute('INSERT INTO departments ("单位名称","科室名称") VALUES (?,?)',
                            ("湘潭市机关事务管理局本级", "办公室"))
            if con.execute('SELECT COUNT(*) FROM users').fetchone()[0] == 0:
                con.execute('INSERT INTO users ("用户名","密码","姓名","所属部门","角色") VALUES (?,?,?,?,?)',
                            ("admin", hash_pw("123456"), "系统管理员", "湘潭市机关事务管理局本级-办公室", "管理员"))
            if con.execute('SELECT COUNT(*) FROM settings').fetchone()[0] == 0:
                con.execute('INSERT INTO settings ("配置项","配置值") VALUES (?,?)', ("default_remind_days", "60"))
            con.commit()
        finally:
            con.close()

# ---- Attachments ----
def read_attachments():
    return db_query("SELECT * FROM attachments ORDER BY rowid")

def write_attachments(rows):
    with db_lock:
        _backup()
        con = _conn()
        try:
            con.execute("DELETE FROM attachments")
            for r in rows:
                _insert_dict(con, "attachments", ATTACH_HEADERS, r)
            con.commit()
        except Exception:
            con.rollback(); raise
        finally:
            con.close()

def attachment_json(row):
    tlabel = row.get("附件类型","")
    tkey = next((k for k,v in ATTACH_TYPES.items() if v==tlabel), "")
    return {
        "id":row.get("附件ID",""),"contract_number":row.get("合同编号",""),
        "type":tkey,"type_label":tlabel,"display_name":row.get("显示名称",""),
        "size":row.get("文件大小",""),"uploader":row.get("上传人",""),
        "uploaded_at":row.get("上传时间",""),
    }

def attach_map_by_number():
    m={}
    for r in read_attachments():
        m.setdefault(str(r.get("合同编号","")),[]).append(attachment_json(r))
    return m

def remove_contract_attachments(number):
    rows = read_attachments()
    keep=[]
    for r in rows:
        if str(r.get("合同编号",""))==str(number):
            fp = os.path.join(ATTACH_DIR, r.get("存储文件名",""))
            try:
                if os.path.exists(fp): os.remove(fp)
            except Exception: pass
        else: keep.append(r)
    if len(keep)!=len(rows): write_attachments(keep)

def read_contracts():
    rows = db_query("SELECT * FROM contracts ORDER BY id")
    result = []
    for r in rows:
        item = {"_row": r.get("id")}
        for cn in CONTRACT_COLUMNS:
            v = r.get(cn, "")
            item[cn] = "" if v is None else v
        for xc in ("是否归档","归档人","归档时间"):
            v2 = r.get(xc, "")
            item[xc] = "" if v2 is None else v2
        result.append(item)
    return result

def c2r(c):
    return {
        "id":f"row_{c['_row']}","number":str(c.get("编号","")),
        "category":c.get("项目分类",""),"name":c.get("项目名称",""),
        "decision_basis":c.get("决策依据",""),"bid_control_price":str(c.get("决策的招标控制价","")),
        "price_basis":c.get("招标控制价决策的依据",""),"procurement_method":c.get("采购方式",""),
        "winner":c.get("中标单位",""),"actual_implementor":c.get("实际实施单位",""),
        "amount":_num(c.get("合同金额（万元）")),
        "fund_source":c.get("资金来源",""),
        "main_content":c.get("合同主要内容",""),"procurement_date":c.get("采购时间",""),
        "start_date":c.get("开工时间",""),"end_date":c.get("完工时间",""),
        "acceptance_date":c.get("竣工验收时间",""),"site_manager":c.get("现场负责人",""),
        "has_change":c.get("是否有签证或变更",""),"change_detail":c.get("签证或变更情况",""),
        "is_settled":c.get("是否结算",""),"settled_amount":_num(c.get("结算金额")),
        "paid_amount":_num(c.get("已支付")),"unpaid_amount":_num(c.get("未支付")),
        "term":c.get("期限",""),"handler":c.get("经办人",""),
        "remark":c.get("备注",""),"department":c.get("所属部门",""),
        "next_bid_date":c.get("下次招采时间",""),"remind_days":_int(c.get("提醒提前天数")),
        "archived":c.get("是否归档","")=="是","archive_by":c.get("归档人",""),"archive_at":c.get("归档时间",""),
    }

def _num(v):
    if v=="" or v is None: return None
    try: return round(float(v),2)
    except: return None

def _int(v):
    if v=="" or v is None: return None
    try: return int(float(v))
    except: return None
def _pending_unpaid(c):
    # 待支付金额：优先取台账"未支付"（含0）；为空时按 合同金额-已支付 计算（空按0）
    up = _num(c.get("未支付"))
    if up is not None:
        return up
    amt = _num(c.get("合同金额（万元）")) or 0
    paid = _num(c.get("已支付")) or 0
    return round(amt - paid, 2)

def _pending_count(cts):
    # 待支付项目数：未归档 且 未支付金额>0
    n = 0
    for c in cts:
        if c.get("是否归档", "") == "是":
            continue
        if _pending_unpaid(c) > 0:
            n += 1
    return n

def _split_names(s):
    # 经办人姓名列表：兼容中文顿号、顿号、斜杠、分号等分隔符
    if s is None: return []
    return [p.strip() for p in str(s).replace("\u3001", ",").replace("/", ",").replace("\uff1b", ";").split(",") if p.strip()]

def upd_cell(rn,cn,val):
    if cn not in COL_IDX: return
    db_exec('UPDATE contracts SET "' + cn + '"=? WHERE id=?', (val, rn))

def upd_row(rn,d):
    sets = [cn for cn in CONTRACT_COLUMNS if cn in d]
    if not sets: return
    params = [d[cn] for cn in sets] + [rn]
    sql = "UPDATE contracts SET " + ",".join('"'+cn+'"=?' for cn in sets) + " WHERE id=?"
    db_exec(sql, params)

def del_row(rn):
    db_exec("DELETE FROM contracts WHERE id=?", (rn,))

def app_row(d):
    cols = list(CONTRACT_COLUMNS)
    with db_lock:
        _backup()
        con = _conn()
        try:
            sql = "INSERT INTO contracts (" + ",".join('"'+c+'"' for c in cols) + ") VALUES (" + ",".join("?"*len(cols)) + ")"
            cur = con.execute(sql, [d.get(cn, "") for cn in cols])
            con.commit()
            return cur.lastrowid
        except Exception:
            con.rollback(); raise
        finally:
            con.close()

# ──── Helpers ────
def get_all_users(): return db_query('SELECT * FROM users')
def get_user(un):
    for u in get_all_users():
        if u["\u7528\u6237\u540d"] == un: return u
    return None

LEADER_DEPT_NAMES = ("\u5c40\u9886\u5bfc", "\u4e3b\u4efb\u73ed\u5b50", "\u4e2d\u5fc3\u73ed\u5b50")
def is_leadership_dept(dept):
    return str(dept or "").strip().split("-")[-1] in LEADER_DEPT_NAMES

def norm_managed(departments, allowed=None):
    if allowed is None: allowed = get_dept_strings()
    if isinstance(departments, (list, tuple)):
        parts = [str(x).strip() for x in departments if str(x).strip()]
    else:
        parts = [p.strip() for p in str(departments or "").replace("\u3001", ",").replace("\uff1b", ";").split(",") if p.strip()]
    bad = [p for p in parts if p not in allowed]
    if bad: raise ValueError("\u5206\u7ba1\u90e8\u95e8\u4e0d\u5b58\u5728\uff1a" + "\u3001".join(bad))
    return "\u3001".join(parts)
def get_all_depts(): return db_query('SELECT * FROM departments')
def get_dept_strings(): return [f"{d['\u5355\u4f4d\u540d\u79f0']}-{d['\u79d1\u5ba4\u540d\u79f0']}" for d in get_all_depts()]
def get_cfg(k):
    rows = db_query('SELECT "配置值" FROM settings WHERE "配置项"=?', (k,))
    return rows[0]["配置值"] if rows else None
def set_cfg(k,v):
    db_exec('INSERT INTO settings ("配置项","配置值") VALUES (?,?) '
            'ON CONFLICT("配置项") DO UPDATE SET "配置值"=excluded."配置值"', (k, str(v)))
def hash_pw(p): return hashlib.sha256((p+SALT).encode()).hexdigest()
def parse_date(s):
    if not s or str(s).strip()=="": return None
    ds = str(s).strip().replace("/","-")
    try:
        p=ds.split("-"); return datetime.date(int(p[0]),int(p[1]),int(p[2]))
    except: return None
def get_default_days():
    v = get_cfg("default_remind_days")
    try: return int(v)
    except: return 60

def get_attach_max_mb():
    v = get_cfg("default_attach_max_mb")
    try:
        n = int(float(v))
        if n < 1: return DEFAULT_ATTACH_MAX_MB
        return n
    except Exception:
        return DEFAULT_ATTACH_MAX_MB

def get_log_keep_days():
    v = get_cfg("log_keep_days")
    try:
        n = int(v)
        if n < 1: return LOG_KEEP_DAYS
        return n
    except Exception:
        return LOG_KEEP_DAYS
def get_cur_user():
    un = session.get("username")
    if not un: return None
    return get_user(un)

# ---- 操作日志（审计） ----
LOG_KEEP_DAYS = 90
_last_log_cleanup = {"date": ""}

def _log_exec(sql, params=()):
    con = _conn()
    try:
        con.execute(sql, params)
        con.commit()
    finally:
        con.close()

def _client_ip():
    ip = request.headers.get("X-Forwarded-For", "") or ""
    if ip and "," in ip: ip = ip.split(",")[0].strip()
    return ip or (request.remote_addr or "")

def _cleanup_logs(days=None):
    if days is None: days = get_log_keep_days()
    cutoff = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
    try:
        _log_exec('DELETE FROM operation_logs WHERE "时间" < ?', (cutoff,))
    except Exception:
        pass

def _maybe_cleanup_logs():
    today = datetime.date.today().isoformat()
    if _last_log_cleanup.get("date") == today: return
    _last_log_cleanup["date"] = today
    _cleanup_logs()

def write_log(module, action, obj="", detail="", result="成功"):
    try:
        u = get_cur_user()
        un = u["用户名"] if u else (session.get("username", "") or "")
        nm = u["姓名"] if u else ""
        rl = u["角色"] if u else ""
    except Exception:
        un = session.get("username", "") or ""; nm = ""; rl = ""
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        _log_exec('INSERT INTO operation_logs ("时间","用户名","姓名","角色","IP","模块","操作类型","对象","详情","结果") VALUES (?,?,?,?,?,?,?,?,?,?)',
                  (ts, un, nm, rl, _client_ip(), module, action, str(obj), str(detail), result))
        _maybe_cleanup_logs()
    except Exception:
        pass

def _visible_depts(u=None):
    # 可见部门集合：本部门 + 分管部门；管理员返回 None 表示全部可见
    if u is None: u = get_cur_user()
    if not u or u.get("\u89d2\u8272") == "\u7ba1\u7406\u5458": return None
    depts = [str(u.get("\u6240\u5c5e\u90e8\u95e8", "")).strip()]
    md = str(u.get("\u5206\u7ba1\u90e8\u95e8", "") or "").strip()
    depts += [p.strip() for p in md.replace("\u3001", ",").replace("\uff1b", ";").split(",") if p.strip()]
    return set(depts)

def _scope_contracts(cts):
    u = get_cur_user()
    if u:
        vis = _visible_depts(u)
        if vis is not None:
            cts = [c for c in cts if str(c.get("\u6240\u5c5e\u90e8\u95e8", "")).strip() in vis]
    return cts

def _can_write(c):
    # 写权限：管理员全部；普通用户限可见部门（本部门+分管部门）
    u = get_cur_user()
    if not u or u["\u89d2\u8272"]=="\u7ba1\u7406\u5458": return True
    vis = _visible_depts(u)
    return vis is not None and str(c.get("\u6240\u5c5e\u90e8\u95e8", "")).strip() in vis

# ──── Auth Decorators ────
def login_required(f):
    @wraps(f)
    def wrapper(*a,**kw):
        if "username" not in session: return jsonify({"error":"\u672a\u767b\u5f55"}),401
        return f(*a,**kw)
    return wrapper

def admin_required(f):
    @wraps(f)
    def wrapper(*a,**kw):
        if "username" not in session: return jsonify({"error":"\u672a\u767b\u5f55"}),401
        u = get_user(session["username"])
        if not u or u.get("\u89d2\u8272")!="\u7ba1\u7406\u5458": return jsonify({"error":"\u6743\u9650\u4e0d\u8db3"}),403
        return f(*a,**kw)
    return wrapper

# ═══════════════════ API Routes ═══════════════════

# ─── Static ───
@app.route("/static/<path:fn>")
def serve_static(fn): return send_from_directory(STATIC_DIR,fn)

# ─── Auth ───
@app.route("/api/login",methods=["POST"])
def api_login():
    d = request.get_json(silent=True) or {}
    un = d.get("username","").strip()
    pw = d.get("password","").strip()
    if not un or not pw: return jsonify({"error":"\u7528\u6237\u540d\u548c\u5bc6\u7801\u4e0d\u80fd\u4e3a\u7a7a"}),400
    u = get_user(un)
    if not u:
        write_log("用户登录", "登录", obj=un, result="失败")
        return jsonify({"error":"用户名或密码错误"}),401
    stored_pw = u.get("密码哈希", u.get("密码", ""))
    if stored_pw != hash_pw(pw):
        write_log("用户登录", "登录", obj=un, result="失败")
        return jsonify({"error":"用户名或密码错误"}),401
    session["username"]=un; session.permanent=True
    write_log("用户登录", "登录", obj=un, result="成功")
    return jsonify({"username":u["\u7528\u6237\u540d"],"name":u["\u59d3\u540d"],"department":u["\u6240\u5c5e\u90e8\u95e8"],"role":u["\u89d2\u8272"]})

@app.route("/api/logout")
def api_logout():
    un = session.get("username","")
    write_log("用户登录", "登出", obj=un)
    session.clear()
    return jsonify({"message":"已退出"})

@app.route("/api/current_user")
@login_required
def api_cur_user():
    u = get_cur_user()
    if not u: session.clear(); return jsonify({"error":"\u7528\u6237\u4e0d\u5b58\u5728"}),401
    return jsonify({"username":u["\u7528\u6237\u540d"],"name":u["\u59d3\u540d"],"department":u["\u6240\u5c5e\u90e8\u95e8"],"role":u["\u89d2\u8272"]})

# ─── Dashboard & Reminders ───
@app.route("/api/dashboard")
@login_required
def api_dashboard():
    u = get_cur_user(); cts = _scope_contracts(read_contracts()); total = len(cts)
    exp = _reminders(u,cts)
    total_amt = sum((_num(c.get("合同金额（万元）")) or 0) for c in cts)
    unpaid_total = sum(_pending_unpaid(c) for c in cts)
    pending_count = _pending_count(cts)
    return jsonify({"total_contracts":total,"total_amount":round(total_amt,2),"dept_contracts":total,"expiring_contracts":len(exp),"pending_project_count":pending_count,"unpaid_amount_total":round(unpaid_total,2)})

def _reminders(u,cts=None):
    if cts is None: cts = read_contracts()
    today = datetime.date.today(); dd = get_default_days(); res = []
    for c in cts:
        nb = parse_date(c.get("\u4e0b\u6b21\u62db\u91c7\u65f6\u95f4",""))
        if not nb: continue
        rd = _int(c.get("\u63d0\u9192\u63d0\u524d\u5929\u6570")); 
        if rd is None: rd = dd
        if today < nb - datetime.timedelta(days=rd): continue
        if u["\u89d2\u8272"]!="\u7ba1\u7406\u5458":
            vis = _visible_depts(u)
            if vis is not None and str(c.get("\u6240\u5c5e\u90e8\u95e8","")).strip() not in vis: continue
        item = c2r(c); item["remaining_days"] = (nb-today).days; res.append(item)
    return res

@app.route("/api/reminders")
@login_required
def api_reminders():
    items=_reminders(get_cur_user()); return jsonify({"count":len(items),"items":items})
@app.route("/api/pending_payments")
@login_required
def api_pending_payments():
    cts = _scope_contracts(read_contracts())
    items = []
    for c in cts:
        # 与“待支付项目数（个）”卡片口径一致：未归档 且 待支付金额>0
        if c.get("是否归档", "") == "是":
            continue
        amt = _pending_unpaid(c)
        if amt <= 0:
            continue
        it = c2r(c)
        it["pending_amount"] = amt
        items.append(it)
    return jsonify({"count": len(items), "items": items})

# ─── Contracts CRUD ───
@app.route("/api/contracts",methods=["GET"])
@login_required
def api_list_contracts():
    cts = _scope_contracts(read_contracts())
    params = {k:request.args.get(k,"").strip() for k in ["unit","dept","person","name","no","category","method","amount_min","amount_max","start_date","end_date","remind_start","remind_end","is_settled","vendor","archived","fund_source"]}
    flt = []
    for c in cts:
        if params["unit"] and params["unit"] not in c.get("\u6240\u5c5e\u90e8\u95e8",""): continue
        if params["dept"] and params["dept"] not in c.get("\u6240\u5c5e\u90e8\u95e8",""): continue
        if params["person"] and params["person"] not in c.get("\u7ecf\u529e\u4eba",""): continue
        if params["name"] and params["name"] not in c.get("\u9879\u76ee\u540d\u79f0",""): continue
        if params["no"] and params["no"] not in c.get("\u7f16\u53f7",""): continue
        if params["category"] and c.get("\u9879\u76ee\u5206\u7c7b","")!=params["category"]: continue
        if params["method"] and c.get("\u91c7\u8d2d\u65b9\u5f0f","")!=params["method"]: continue
        if params["fund_source"] and c.get("\u8d44\u91d1\u6765\u6e90","")!=params["fund_source"]: continue
        amt = _num(c.get("\u5408\u540c\u91d1\u989d\uff08\u4e07\u5143\uff09"))
        if params["amount_min"]:
            try:
                if amt is None or amt<float(params["amount_min"]): continue
            except: pass
        if params["amount_max"]:
            try:
                if amt is None or amt>float(params["amount_max"]): continue
            except: pass
        cd = parse_date(c.get("\u91c7\u8d2d\u65f6\u95f4",""))
        if params["start_date"]:
            sd=parse_date(params["start_date"])
            if sd and (cd is None or cd<sd): continue
        if params["end_date"]:
            ed=parse_date(params["end_date"])
            if ed and (cd is None or cd>ed): continue
        nd = parse_date(c.get("\u4e0b\u6b21\u62db\u91c7\u65f6\u95f4",""))
        if params["remind_start"]:
            rs=parse_date(params["remind_start"])
            if rs and (nd is None or nd<rs): continue
        if params["remind_end"]:
            re_=parse_date(params["remind_end"])
            if re_ and (nd is None or nd>re_): continue
        if params["is_settled"] in ("\u662f","\u5426") and c.get("\u662f\u5426\u7ed3\u7b97","")!=params["is_settled"]: continue
        if params["vendor"]:
            if params["vendor"] not in c.get("\u4e2d\u6807\u5355\u4f4d","") and params["vendor"] not in c.get("\u5b9e\u9645\u5b9e\u65bd\u5355\u4f4d",""): continue
        if params["archived"] in ("是","否") and (c.get("是否归档","") or "否") != params["archived"]: continue
        flt.append(c)
    # 列表默认按“编号”降序排列（数字编号按数值降序，非数字编号排后）
    def _no_key(c):
        v = str(c.get("编号", "") or "").strip()
        try:
            return (0, -float(v))
        except Exception:
            return (1, v)
    flt.sort(key=_no_key)
    try: p=int(request.args.get("page",1)); ps=min(int(request.args.get("page_size",20)),200)
    except: p=1; ps=20
    si=(p-1)*ps; ei=si+ps
    att_map = attach_map_by_number()
    items=[]
    for x in flt[si:ei]:
        it = c2r(x)
        it["attachments"] = att_map.get(str(x.get("\u7f16\u53f7","")), [])
        items.append(it)
    return jsonify({"total":len(flt),"page":p,"page_size":ps,"items":items})

@app.route("/api/contracts",methods=["POST"])
@login_required
def api_create_contract():
    u = get_cur_user()
    if u["\u89d2\u8272"]=="\u67e5\u8be2\u7528\u6237": return jsonify({"error":"\u6743\u9650\u4e0d\u8db3"}),403
    d = request.get_json(silent=True) or {}
    for f in ["number","category","name","procurement_method","amount","main_content","handler","fund_source"]:
        if not d.get(f,""): return jsonify({"error":f"\u5b57\u6bb5 '{f}' \u4e0d\u80fd\u4e3a\u7a7a"}),400
    for c in read_contracts():
        if c.get("\u7f16\u53f7","")==d["number"]: return jsonify({"error":f"\u7f16\u53f7 '{d['number']}' \u5df2\u5b58\u5728"}),400
    rd = {cn:"" for cn in CONTRACT_COLUMNS}
    fm = {"number":"\u7f16\u53f7","category":"\u9879\u76ee\u5206\u7c7b","name":"\u9879\u76ee\u540d\u79f0","decision_basis":"\u51b3\u7b56\u4f9d\u636e","bid_control_price":"\u51b3\u7b56\u7684\u62db\u6807\u63a7\u5236\u4ef7","price_basis":"\u62db\u6807\u63a7\u5236\u4ef7\u51b3\u7b56\u7684\u4f9d\u636e","procurement_method":"\u91c7\u8d2d\u65b9\u5f0f","winner":"\u4e2d\u6807\u5355\u4f4d","actual_implementor":"\u5b9e\u9645\u5b9e\u65bd\u5355\u4f4d","amount":"\u5408\u540c\u91d1\u989d\uff08\u4e07\u5143\uff09","fund_source":"\u8d44\u91d1\u6765\u6e90","main_content":"\u5408\u540c\u4e3b\u8981\u5185\u5bb9","procurement_date":"\u91c7\u8d2d\u65f6\u95f4","start_date":"\u5f00\u5de5\u65f6\u95f4","end_date":"\u5b8c\u5de5\u65f6\u95f4","acceptance_date":"\u7ae3\u5de5\u9a8c\u6536\u65f6\u95f4","site_manager":"\u73b0\u573a\u8d1f\u8d23\u4eba","has_change":"\u662f\u5426\u6709\u7b7e\u8bc1\u6216\u53d8\u66f4","change_detail":"\u7b7e\u8bc1\u6216\u53d8\u66f4\u60c5\u51b5","is_settled":"\u662f\u5426\u7ed3\u7b97","settled_amount":"\u7ed3\u7b97\u91d1\u989d","paid_amount":"\u5df2\u652f\u4ed8","unpaid_amount":"\u672a\u652f\u4ed8","term":"\u671f\u9650","handler":"\u7ecf\u529e\u4eba","remark":"\u5907\u6ce8","next_bid_date":"\u4e0b\u6b21\u62db\u91c7\u65f6\u95f4","remind_days":"\u63d0\u9192\u63d0\u524d\u5929\u6570"}
    for rk,ck in fm.items():
        if rk in d and d[rk] is not None: rd[ck]=d[rk]
    rd["\u6240\u5c5e\u90e8\u95e8"]=u["\u6240\u5c5e\u90e8\u95e8"]
    if not rd.get("\u63d0\u9192\u63d0\u524d\u5929\u6570") or str(rd.get("\u63d0\u9192\u63d0\u524d\u5929\u6570","")).strip()=="":
        rd["\u63d0\u9192\u63d0\u524d\u5929\u6570"]=get_default_days()
    app_row(rd)
    write_log("合同台账", "新增", obj=d["number"], detail=rd.get("\u9879\u76ee\u540d\u79f0",""))
    for c in read_contracts():
        if str(c.get("\u7f16\u53f7",""))==str(d["number"]):
            return jsonify({"message":"\u5408\u540c\u521b\u5efa\u6210\u529f","number":d["number"],"id":f"row_{c['_row']}"}),201
    return jsonify({"message":"\u5408\u540c\u521b\u5efa\u6210\u529f","number":d["number"]}),201

# ─── Export ───
@app.route("/api/contracts/export",methods=["GET"])
@login_required
def api_export():
    cts = _scope_contracts(read_contracts())
    params = {k:request.args.get(k,"").strip() for k in ["unit","dept","person","name","no","category","method","amount_min","amount_max","start_date","end_date","remind_start","remind_end","is_settled","vendor","archived","fund_source"]}
    flt = list(cts)
    if params["unit"]: flt=[c for c in flt if params["unit"] in c.get("\u6240\u5c5e\u90e8\u95e8","")]
    if params["dept"]: flt=[c for c in flt if params["dept"] in c.get("\u6240\u5c5e\u90e8\u95e8","")]
    if params["person"]: flt=[c for c in flt if params["person"] in c.get("\u7ecf\u529e\u4eba","")]
    if params["name"]: flt=[c for c in flt if params["name"] in c.get("\u9879\u76ee\u540d\u79f0","")]
    if params["no"]: flt=[c for c in flt if params["no"] in c.get("\u7f16\u53f7","")]
    if params["category"]: flt=[c for c in flt if c.get("\u9879\u76ee\u5206\u7c7b","")==params["category"]]
    if params["method"]: flt=[c for c in flt if c.get("\u91c7\u8d2d\u65b9\u5f0f","")==params["method"]]
    if params["fund_source"]: flt=[c for c in flt if c.get("\u8d44\u91d1\u6765\u6e90","")==params["fund_source"]]
    if params["amount_min"]:
        try: mn=float(params["amount_min"]); flt=[c for c in flt if _num(c.get("\u5408\u540c\u91d1\u989d\uff08\u4e07\u5143\uff09")) is not None and _num(c.get("\u5408\u540c\u91d1\u989d\uff08\u4e07\u5143\uff09"))>=mn]
        except: pass
    if params["amount_max"]:
        try: mx=float(params["amount_max"]); flt=[c for c in flt if _num(c.get("\u5408\u540c\u91d1\u989d\uff08\u4e07\u5143\uff09")) is not None and _num(c.get("\u5408\u540c\u91d1\u989d\uff08\u4e07\u5143\uff09"))<=mx]
        except: pass
    if params["start_date"]:
        sd=parse_date(params["start_date"])
        if sd: flt=[c for c in flt if parse_date(c.get("\u91c7\u8d2d\u65f6\u95f4","")) and parse_date(c.get("\u91c7\u8d2d\u65f6\u95f4",""))>=sd]
    if params["end_date"]:
        ed=parse_date(params["end_date"])
        if ed: flt=[c for c in flt if parse_date(c.get("\u91c7\u8d2d\u65f6\u95f4","")) and parse_date(c.get("\u91c7\u8d2d\u65f6\u95f4",""))<=ed]
    if params["remind_start"]:
        rs=parse_date(params["remind_start"])
        if rs: flt=[c for c in flt if parse_date(c.get("\u4e0b\u6b21\u62db\u91c7\u65f6\u95f4","")) and parse_date(c.get("\u4e0b\u6b21\u62db\u91c7\u65f6\u95f4",""))>=rs]
    if params["remind_end"]:
        re_=parse_date(params["remind_end"])
        if re_: flt=[c for c in flt if parse_date(c.get("\u4e0b\u6b21\u62db\u91c7\u65f6\u95f4","")) and parse_date(c.get("\u4e0b\u6b21\u62db\u91c7\u65f6\u95f4",""))<=re_]
    if params["is_settled"] in ("\u662f","\u5426"): flt=[c for c in flt if c.get("\u662f\u5426\u7ed3\u7b97","")==params["is_settled"]]
    if params["vendor"]: flt=[c for c in flt if params["vendor"] in c.get("\u4e2d\u6807\u5355\u4f4d","") or params["vendor"] in c.get("\u5b9e\u9645\u5b9e\u65bd\u5355\u4f4d","")]
    if params["archived"] in ("是","否"): flt=[c for c in flt if (c.get("是否归档","") or "否")==params["archived"]]

    wb = openpyxl.Workbook(); ws = wb.active; ws.title="\u5408\u540c\u53f0\u8d26"
    for ci,cn in enumerate(CONTRACT_COLUMNS,1): ws.cell(row=1,column=ci,value=cn).font=Font(bold=True)
    for ri,c in enumerate(flt,2):
        for ci,cn in enumerate(CONTRACT_COLUMNS,1):
            v=c.get(cn,"")
            if isinstance(v,float): v=round(v,2)
            ws.cell(row=ri,column=ci,value=v)
    bio=BytesIO(); wb.save(bio); bio.seek(0)
    ts=datetime.datetime.now().strftime("%Y%m%d")
    from urllib.parse import quote
    fn = f"\u5408\u540c\u53f0\u8d26\u5bfc\u51fa_{ts}.xlsx"
    write_log("合同台账", "导出", obj="合同台账", detail="导出 " + str(len(flt)) + " 条")
    return Response(bio.getvalue(),mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition":f"attachment; filename*=UTF-8''{quote(fn)}"})

# ─── Import ───
IMPORT_ENUM_CHECK = {
    "项目分类": ["服务类", "维护维修类", "货物类", "建设项目"],
    "采购方式": ["公开招标", "竞争性磋商", "竞争性谈判", "自行采购"],
    "资金来源": ["年初部门预算", "年中追加预算", "统筹预算", "上级专项转移支付", "上年结转结余", "其他"],
}
IMPORT_YN_FIELDS = ("是否有签证或变更", "是否结算")
IMPORT_DATE_FIELDS = ("采购时间", "开工时间", "完工时间", "竣工验收时间", "下次招采时间")
IMPORT_NUM_FIELDS = ("决策的招标控制价", "合同金额（万元）", "结算金额", "已支付", "未支付")
TEMPLATE_EXAMPLE = {
    "编号": "2026-FW-001", "项目分类": "服务类", "项目名称": "示例合同（导入前请删除此行）",
    "决策的招标控制价": 10, "采购方式": "自行采购",
    "中标单位": "示例供应商", "实际实施单位": "示例供应商",
    "合同金额（万元）": 9.5, "资金来源": "年初部门预算", "合同主要内容": "示例内容",
    "采购时间": "2026-01-01", "开工时间": "2026-02-01", "完工时间": "2026-12-31",
    "现场负责人": "张三", "是否有签证或变更": "否", "是否结算": "否",
    "已支付": 5, "未支付": 4.5, "期限": "一年", "经办人": "张三",
    "备注": "示例行，导入前请删除",
    "所属部门": "湘潭市机关事务管理局本级-办公室",
    "下次招采时间": "2027-01-01", "提醒提前天数": 60,
}

def _norm_header(cn):
    if cn == "端工验收时间": return "竣工验收时间"
    if cn in ("经办人/分管领导",): return "经办人"
    if cn in ("甲方现场负责人",): return "现场负责人"
    return cn

def _build_import_vals(hdr, src):
    vals = {cn: "" for cn in CONTRACT_COLUMNS}
    for ci, cn in enumerate(hdr):
        cn2 = _norm_header(cn)
        if cn2 in vals and ci < len(src):
            vals[cn2] = _norm(src[ci])
    return vals

def _check_import_row(vals, existing, file_seen, cur_row):
    errors, warns = [], []
    number = str(vals.get("编号", "")).strip()
    name = str(vals.get("项目名称", "")).strip()
    if not number: errors.append("编号为空")
    if not name: errors.append("项目名称为空")
    if number:
        if number in existing: errors.append("编号已存在于数据库")
        if number in file_seen: errors.append("编号在文件内重复（首次出现于第%d行）" % file_seen[number])
        else: file_seen[number] = cur_row
    # 必填字段校验（与数据库 NOT NULL 、前端必填星号一致：编号/项目名称已在上方单独校验）
    IMPORT_REQUIRED = ("项目分类", "采购方式", "合同金额（万元）", "资金来源", "合同主要内容", "经办人", "所属部门")
    for cn in IMPORT_REQUIRED:
        if not str(vals.get(cn, "")).strip():
            errors.append("%s为空（必填）" % cn)
    for cn in IMPORT_NUM_FIELDS:
        v = vals.get(cn, "")
        if v not in ("", None) and _num(v) is None:
            errors.append("%s不是有效数字" % cn)
    for cn in IMPORT_DATE_FIELDS:
        v = vals.get(cn, "")
        if v not in ("", None) and parse_date(v) is None:
            errors.append("%s日期格式应为yyyy-mm-dd" % cn)
    rv = vals.get("提醒提前天数", "")
    if rv not in ("", None) and _int(rv) is None:
        errors.append("提醒提前天数应为整数")
    dept = str(vals.get("所属部门", "")).strip()
    if dept and dept not in get_dept_strings():
        warns.append("所属部门不存在于部门列表")
    for cn, allowed in IMPORT_ENUM_CHECK.items():
        v = str(vals.get(cn, "")).strip()
        if v and v not in allowed:
            warns.append("%s取值不在标准选项内" % cn)
    for cn in IMPORT_YN_FIELDS:
        v = str(vals.get(cn, "")).strip()
        if v and v not in ("是", "否"):
            warns.append("%s应为是或否" % cn)
    if errors: return ("error", errors)
    if warns: return ("warn", warns)
    return ("ok", [])

def _load_import_file():
    f = request.files.get("file")
    if not f or not f.filename:
        return None, None, "请选择要导入的文件"
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return None, None, "仅支持 .xlsx 格式文件"
    try:
        wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True)
    except Exception:
        return None, None, "文件解析失败，请使用有效的 xlsx 文件"
    sheet_name = "合同台账" if "合同台账" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return None, None, "文件中没有可导入的数据"
    hdr = [_norm_header(str(x).strip() if x is not None else "") for x in rows[0]]
    missing = [cn for cn in ("编号", "项目名称") if cn not in hdr]
    if missing:
        return None, None, "文件表头缺少列：" + "、".join(missing) + "，请使用导出的合同台账模板"
    return rows, hdr, None

def _import_existing_numbers():
    return {str(c.get("编号", "")).strip() for c in read_contracts()}

@app.route("/api/contracts/template", methods=["GET"])
@admin_required
def api_contract_template():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "合同台账"
    headers = list(CONTRACT_COLUMNS)
    for ci, cn in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=cn)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E74B5")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ci, cn in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 16
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    # 示例行
    for ci, cn in enumerate(headers, 1):
        v = TEMPLATE_EXAMPLE.get(cn, "")
        if v != "":
            ws.cell(row=2, column=ci, value=v)
    ws.cell(row=2, column=1).font = Font(italic=True, color="999999")
    # 下拉验证
    def add_dv(formula, col_letter):
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("%s3:%s1000" % (col_letter, col_letter))
    col_of = {cn: openpyxl.utils.get_column_letter(i + 1) for i, cn in enumerate(headers)}
    add_dv('"服务类,维护维修类,货物类,建设项目"', col_of["项目分类"])
    add_dv('"公开招标,竞争性磋商,竞争性谈判,自行采购"', col_of["采购方式"])
    add_dv('"是,否"', col_of["是否有签证或变更"])
    add_dv('"是,否"', col_of["是否结算"])
    add_dv('"年初部门预算,年中追加预算,统筹预算,上级专项转移支付,上年结转结余,其他"', col_of["资金来源"])
    # 填写说明
    ws2 = wb.create_sheet("填写说明")
    tips = [
        ("合同台账导入模板填写说明", True),
        ("", False),
        ("1. 第一行为表头，不可删除；第二行为示例，导入前请删除。", False),
        ("2. 必填字段：编号、项目分类、项目名称、采购方式、合同金额（万元）、资金来源、合同主要内容、经办人、所属部门。", False),
        ("3. 编号全局唯一，与数据库重复的行将被跳过不导入。", False),
        ("4. 日期列（采购时间、开工时间、完工时间、竣工验收时间、下次招采时间）格式为 yyyy-mm-dd。", False),
        ("5. 金额列（合同金额、结算金额、已支付、未支付）填写数字。", False),
        ("6. 经办人多人用顿号、分隔；所属部门格式为“单位名称-科室名称”。", False),
        ("7. 下次招采时间填写后系统会按提醒提前天数进行到期提醒；提醒提前天数为空时使用全局默认值。", False),
        ("8. 导入前可先使用“导入预览”查看逐条校验结果，错误行不会被导入。", False),
    ]
    for i, (txt, bold) in enumerate(tips, 1):
        cell = ws2.cell(row=i, column=1, value=txt)
        if bold:
            cell.font = Font(bold=True, size=13, color="1F4D78")
    ws2.column_dimensions["A"].width = 90
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    from urllib.parse import quote
    fn = "合同台账导入模板.xlsx"
    write_log("合同台账", "下载模板", obj="合同台账", detail="下载导入模板")
    return Response(bio.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename*=UTF-8''%s" % quote(fn)})

@app.route("/api/contracts/import/preview", methods=["POST"])
@admin_required
def api_import_preview():
    rows, hdr, err = _load_import_file()
    if err: return jsonify({"error": err}), 400
    existing = _import_existing_numbers()
    file_seen = {}
    items = []
    for ri in range(1, len(rows)):
        src = rows[ri]
        if src is None: continue
        if all(v is None or str(v).strip() == "" for v in src): continue
        vals = _build_import_vals(hdr, src)
        number = str(vals.get("编号", "")).strip()
        name = str(vals.get("项目名称", "")).strip()
        status, reasons = _check_import_row(vals, existing, file_seen, ri + 1)
        items.append({"row": ri + 1, "number": number, "name": name, "status": status, "reasons": reasons})
    ok_count = sum(1 for it in items if it["status"] == "ok")
    warn_count = sum(1 for it in items if it["status"] == "warn")
    error_count = sum(1 for it in items if it["status"] == "error")
    return jsonify({
        "filename": request.files.get("file").filename,
        "total": len(items), "ok_count": ok_count, "warn_count": warn_count, "error_count": error_count,
        "importable": ok_count + warn_count, "items": items,
    })

@app.route("/api/contracts/import", methods=["POST"])
@admin_required
def api_import_contracts():
    rows, hdr, err = _load_import_file()
    if err: return jsonify({"error": err}), 400
    existing = _import_existing_numbers()
    file_seen = {}
    to_insert = []
    skipped = []
    for ri in range(1, len(rows)):
        src = rows[ri]
        if src is None: continue
        if all(v is None or str(v).strip() == "" for v in src): continue
        vals = _build_import_vals(hdr, src)
        number = str(vals.get("编号", "")).strip()
        name = str(vals.get("项目名称", "")).strip()
        status, reasons = _check_import_row(vals, existing, file_seen, ri + 1)
        if status == "error":
            skipped.append({"row": ri + 1, "number": number, "name": name, "reasons": reasons})
            continue
        to_insert.append(vals)
        existing.add(number)
    added = 0
    if to_insert:
        with db_lock:
            _backup()
            con = _conn()
            try:
                cols = list(CONTRACT_COLUMNS)
                sql = "INSERT INTO contracts (" + ",".join('"'+c+'"' for c in cols) + ") VALUES (" + ",".join("?"*len(cols)) + ")"
                for vals in to_insert:
                    con.execute(sql, [vals[cn] for cn in cols])
                    added += 1
                con.commit()
            except sqlite3.IntegrityError as e:
                con.rollback()
                return jsonify({"error": "数据库约束拒绝写入（必填字段为空、编号重复、金额非负、日期格式等）：%s" % e}), 400
            except Exception:
                con.rollback(); raise
            finally:
                con.close()
    msg = "导入完成：新增 %d 条" % added
    if skipped: msg += "，跳过 %d 条" % len(skipped)
    write_log("合同台账", "导入", obj=str(request.files.get("file").filename or ""), detail=msg)
    return jsonify({"message": msg, "added": added, "skipped": skipped}), 200

# ─── Get Single Contract ───
@app.route("/api/contracts/<cid>",methods=["GET"])
@login_required
def api_get_contract(cid):
    u = get_cur_user()
    if not cid.startswith("row_"): return jsonify({"error":"\u65e0\u6548\u7684\u5408\u540cID"}),400
    try: rn=int(cid.split("_")[1])
    except: return jsonify({"error":"\u65e0\u6548\u7684\u5408\u540cID"}),400
    for c in read_contracts():
        if c["_row"]==rn:
            if u["\u89d2\u8272"]!="\u7ba1\u7406\u5458":
                vis = _visible_depts(u)
                if vis is not None and str(c.get("\u6240\u5c5e\u90e8\u95e8","")).strip() not in vis:
                    return jsonify({"error":"\u5408\u540c\u4e0d\u5b58\u5728"}),404
            it = c2r(c)
            it["attachments"] = attach_map_by_number().get(str(c.get("\u7f16\u53f7","")), [])
            return jsonify(it)
    return jsonify({"error":"\u5408\u540c\u4e0d\u5b58\u5728"}),404

# ─── Update ───
@app.route("/api/contracts/<cid>",methods=["PUT"])
@login_required
def api_update_contract(cid):
    u = get_cur_user()
    if u["\u89d2\u8272"]=="\u67e5\u8be2\u7528\u6237": return jsonify({"error":"\u6743\u9650\u4e0d\u8db3"}),403
    if not cid.startswith("row_"): return jsonify({"error":"\u65e0\u6548\u7684\u5408\u540cID"}),400
    try: rn=int(cid.split("_")[1])
    except: return jsonify({"error":"\u65e0\u6548\u7684\u5408\u540cID"}),400
    t=None
    for c in read_contracts():
        if c["_row"]==rn: t=c; break
    if not t: return jsonify({"error":"\u5408\u540c\u4e0d\u5b58\u5728"}),404
    if str(t.get("是否归档",""))=="是": return jsonify({"error":"合同已归档，仅可浏览，不可修改"}),403
    if not _can_write(t): return jsonify({"error":"\u53ea\u80fd\u7f16\u8f91\u672c\u90e8\u95e8\u6216\u5206\u7ba1\u90e8\u95e8\u5408\u540c"}),403
    d=request.get_json(silent=True) or {}
    if "fund_source" in d and str(d.get("fund_source","")).strip()=="": return jsonify({"error":"字段 'fund_source' 不能为空"}),400
    ud={}
    fm={"number":"\u7f16\u53f7","category":"\u9879\u76ee\u5206\u7c7b","name":"\u9879\u76ee\u540d\u79f0","decision_basis":"\u51b3\u7b56\u4f9d\u636e","bid_control_price":"\u51b3\u7b56\u7684\u62db\u6807\u63a7\u5236\u4ef7","price_basis":"\u62db\u6807\u63a7\u5236\u4ef7\u51b3\u7b56\u7684\u4f9d\u636e","procurement_method":"\u91c7\u8d2d\u65b9\u5f0f","winner":"\u4e2d\u6807\u5355\u4f4d","actual_implementor":"\u5b9e\u9645\u5b9e\u65bd\u5355\u4f4d","amount":"\u5408\u540c\u91d1\u989d\uff08\u4e07\u5143\uff09","fund_source":"\u8d44\u91d1\u6765\u6e90","main_content":"\u5408\u540c\u4e3b\u8981\u5185\u5bb9","procurement_date":"\u91c7\u8d2d\u65f6\u95f4","start_date":"\u5f00\u5de5\u65f6\u95f4","end_date":"\u5b8c\u5de5\u65f6\u95f4","acceptance_date":"\u7ae3\u5de5\u9a8c\u6536\u65f6\u95f4","site_manager":"\u73b0\u573a\u8d1f\u8d23\u4eba","has_change":"\u662f\u5426\u6709\u7b7e\u8bc1\u6216\u53d8\u66f4","change_detail":"\u7b7e\u8bc1\u6216\u53d8\u66f4\u60c5\u51b5","is_settled":"\u662f\u5426\u7ed3\u7b97","settled_amount":"\u7ed3\u7b97\u91d1\u989d","paid_amount":"\u5df2\u652f\u4ed8","unpaid_amount":"\u672a\u652f\u4ed8","term":"\u671f\u9650","handler":"\u7ecf\u529e\u4eba","remark":"\u5907\u6ce8","next_bid_date":"\u4e0b\u6b21\u62db\u91c7\u65f6\u95f4","remind_days":"\u63d0\u9192\u63d0\u524d\u5929\u6570"}
    for rk,ck in fm.items():
        if rk in d and d[rk] is not None: ud[ck]=d[rk]
    old_num = str(t.get("\u7f16\u53f7",""))
    new_num = ud.get("\u7f16\u53f7")
    if new_num is not None and str(new_num)!=old_num:
        for c in read_contracts():
            if c["_row"]!=rn and str(c.get("\u7f16\u53f7",""))==str(new_num):
                return jsonify({"error":f"\u7f16\u53f7 '{new_num}' \u5df2\u5b58\u5728"}),400
        rows = read_attachments(); changed=False
        for r in rows:
            if str(r.get("\u5408\u540c\u7f16\u53f7",""))==old_num: r["\u5408\u540c\u7f16\u53f7"]=str(new_num); changed=True
        if changed: write_attachments(rows)
    upd_row(rn,ud)
    write_log("合同台账", "修改", obj=str(t.get("\u7f16\u53f7","")))
    return jsonify({"message":"\u5408\u540c\u66f4\u65b0\u6210\u529f"})

# ─── Delete ───
@app.route("/api/contracts/<cid>",methods=["DELETE"])
@login_required
def api_delete_contract(cid):
    u=get_cur_user()
    if u["\u89d2\u8272"]=="\u67e5\u8be2\u7528\u6237": return jsonify({"error":"\u6743\u9650\u4e0d\u8db3"}),403
    if not cid.startswith("row_"): return jsonify({"error":"\u65e0\u6548\u7684\u5408\u540cID"}),400
    try: rn=int(cid.split("_")[1])
    except: return jsonify({"error":"\u65e0\u6548\u7684\u5408\u540cID"}),400
    t=None
    for c in read_contracts():
        if c["_row"]==rn: t=c; break
    if not t: return jsonify({"error":"\u5408\u540c\u4e0d\u5b58\u5728"}),404
    if str(t.get("是否归档",""))=="是": return jsonify({"error":"合同已归档，仅可浏览，不可修改"}),403
    if not _can_write(t): return jsonify({"error":"\u53ea\u80fd\u5220\u9664\u672c\u90e8\u95e8\u6216\u5206\u7ba1\u90e8\u95e8\u5408\u540c"}),403
    data = request.get_json(silent=True) or {}
    reason = str(data.get("reason","")).strip()
    if not reason: return jsonify({"error":"请填写删除原因"}),400

    remove_contract_attachments(t.get("\u7f16\u53f7",""))
    del_row(rn)
    write_log("合同台账", "删除", obj=str(t.get("\u7f16\u53f7","")), detail="删除原因：" + reason)
    return jsonify({"message":"\u5408\u540c\u5220\u9664\u6210\u529f"})

# ---- Archive ----
@app.route("/api/contracts/<cid>/archive",methods=["PUT"])
@login_required
def api_archive_contract(cid):
    u=get_cur_user()
    if u["角色"]=="查询用户": return jsonify({"error":"权限不足"}),403
    if not cid.startswith("row_"): return jsonify({"error":"无效的合同ID"}),400
    try: rn=int(cid.split("_")[1])
    except: return jsonify({"error":"无效的合同ID"}),400
    t=None
    for c in read_contracts():
        if c["_row"]==rn: t=c; break
    if not t: return jsonify({"error":"合同不存在"}),404
    if str(t.get("是否归档",""))=="是": return jsonify({"message":"该合同已归档"}),200
    if u["角色"]!="管理员":
        pf = str(t.get("经办人",""))
        names = _split_names(pf)
        if u["姓名"] not in names:
            return jsonify({"error":"仅管理员或该合同经办人可执行归档"}),403
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    db_exec('UPDATE contracts SET "是否归档"=?, "归档人"=?, "归档时间"=? WHERE id=?',
            ("是", u.get("姓名", u.get("用户名","")), ts, rn))
    write_log("合同台账", "归档", obj=str(t.get("编号","")))
    return jsonify({"message":"合同已归档，进入只读状态"})

# ---- Attachments ----
@app.route("/api/contracts/<cid>/attachments",methods=["POST"])
@login_required
def api_upload_attachment(cid):
    u = get_cur_user()
    if u["\u89d2\u8272"]=="\u67e5\u8be2\u7528\u6237": return jsonify({"error":"\u6743\u9650\u4e0d\u8db3"}),403
    if not cid.startswith("row_"): return jsonify({"error":"\u65e0\u6548\u7684\u5408\u540cID"}),400
    try: rn=int(cid.split("_")[1])
    except: return jsonify({"error":"\u65e0\u6548\u7684\u5408\u540cID"}),400
    t=None
    for c in read_contracts():
        if c["_row"]==rn: t=c; break
    if not t: return jsonify({"error":"\u5408\u540c\u4e0d\u5b58\u5728"}),404
    if str(t.get("是否归档",""))=="是": return jsonify({"error":"合同已归档，仅可浏览，不可修改"}),403
    if not _can_write(t): return jsonify({"error":"\u53ea\u80fd\u4e3a\u672c\u90e8\u95e8\u6216\u5206\u7ba1\u90e8\u95e8\u5408\u540c\u4e0a\u4f20\u9644\u4ef6"}),403
    atype = (request.form.get("type","") or "").strip()
    if atype not in ATTACH_TYPES: return jsonify({"error":"\u9644\u4ef6\u7c7b\u578b\u65e0\u6548"}),400
    f = request.files.get("file")
    if not f or not f.filename: return jsonify({"error":"\u8bf7\u9009\u62e9\u8981\u4e0a\u4f20\u7684\u6587\u4ef6"}),400
    display = os.path.basename(f.filename.replace("\\","/")).strip()[:200] or "\u9644\u4ef6"
    ext = os.path.splitext(display)[1].lower()
    if ext not in ATTACH_ALLOWED_EXT:
        return jsonify({"error":"\u4ec5\u652f\u6301\u56fe\u7247\u6216PDF\u683c\u5f0f\uff08jpg/jpeg/png/gif/webp/bmp/pdf\uff09"}),400
    max_mb = get_attach_max_mb()
    max_bytes = max_mb * 1024 * 1024
    store_name = uuid.uuid4().hex + ext
    os.makedirs(ATTACH_DIR,exist_ok=True)
    fp = os.path.join(ATTACH_DIR, store_name)
    f.save(fp)
    size = os.path.getsize(fp)
    if size > max_bytes:
        try: os.remove(fp)
        except Exception: pass
        return jsonify({"error":"附件大小不能超过 %dM（当前 %.2fM）" % (max_mb, size/1048576.0)}),400
    rows = read_attachments()
    rows.append({
        "\u9644\u4ef6ID": uuid.uuid4().hex,
        "\u5408\u540c\u7f16\u53f7": str(t.get("\u7f16\u53f7","")),
        "\u9644\u4ef6\u7c7b\u578b": ATTACH_TYPES[atype],
        "\u663e\u793a\u540d\u79f0": display,
        "\u5b58\u50a8\u6587\u4ef6\u540d": store_name,
        "\u6587\u4ef6\u5927\u5c0f": str(size),
        "\u4e0a\u4f20\u4eba": u.get("\u59d3\u540d", u.get("\u7528\u6237\u540d","")),
        "\u4e0a\u4f20\u65f6\u95f4": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })
    write_attachments(rows)
    write_log("合同台账", "上传附件", obj=str(t.get("\u7f16\u53f7","")), detail=display)
    return jsonify({"message":"\u9644\u4ef6\u4e0a\u4f20\u6210\u529f","attachment":attachment_json(rows[-1])}),201

@app.route("/api/attachments/<att_id>/download",methods=["GET"])
@login_required
def api_download_attachment(att_id):
    u = get_cur_user()
    row = next((r for r in read_attachments() if r.get("\u9644\u4ef6ID","")==att_id), None)
    if not row: return jsonify({"error":"\u9644\u4ef6\u4e0d\u5b58\u5728"}),404
    if u["\u89d2\u8272"]!="\u7ba1\u7406\u5458":
        vis = _visible_depts(u)
        cnum = str(row.get("\u5408\u540c\u7f16\u53f7",""))
        if vis is not None and not any(str(c.get("\u7f16\u53f7",""))==cnum and str(c.get("\u6240\u5c5e\u90e8\u95e8","")).strip() in vis for c in read_contracts()):
            return jsonify({"error":"\u6743\u9650\u4e0d\u8db3"}),403
    fp = os.path.join(ATTACH_DIR, row.get("\u5b58\u50a8\u6587\u4ef6\u540d",""))
    if not os.path.exists(fp): return jsonify({"error":"\u9644\u4ef6\u6587\u4ef6\u4e0d\u5b58\u5728"}),404
    return send_file(fp, as_attachment=True, download_name=row.get("\u663e\u793a\u540d\u79f0","\u9644\u4ef6"))

@app.route("/api/attachments/<att_id>",methods=["DELETE"])
@login_required
def api_delete_attachment(att_id):
    u=get_cur_user()
    if u["\u89d2\u8272"]=="\u67e5\u8be2\u7528\u6237": return jsonify({"error":"\u6743\u9650\u4e0d\u8db3"}),403
    rows = read_attachments()
    row = next((r for r in rows if r.get("\u9644\u4ef6ID","")==att_id), None)
    if not row: return jsonify({"error":"\u9644\u4ef6\u4e0d\u5b58\u5728"}),404
    for c in read_contracts():
        if str(c.get("\u7f16\u53f7",""))==str(row.get("\u5408\u540c\u7f16\u53f7","")):
            if not _can_write(c): return jsonify({"error":"\u6743\u9650\u4e0d\u8db3"}),403
            if c.get("是否归档","")=="是": return jsonify({"error":"合同已归档，仅可浏览，不可修改"}),403
            break
    fp = os.path.join(ATTACH_DIR, row.get("\u5b58\u50a8\u6587\u4ef6\u540d",""))
    try:
        if os.path.exists(fp): os.remove(fp)
    except Exception: pass
    write_attachments([r for r in rows if r.get("\u9644\u4ef6ID","")!=att_id])
    write_log("合同台账", "删除附件", obj=str(row.get("\u5408\u540c\u7f16\u53f7","")), detail=str(row.get("\u663e\u793a\u540d\u79f0","")))
    return jsonify({"message":"\u9644\u4ef6\u5df2\u5220\u9664"})

# ─── Clear Remind ───
@app.route("/api/contracts/<cid>/clear_remind",methods=["PUT"])
@login_required
def api_clear_remind(cid):
    u=get_cur_user()
    if u["\u89d2\u8272"]=="\u67e5\u8be2\u7528\u6237": return jsonify({"error":"\u6743\u9650\u4e0d\u8db3"}),403
    if not cid.startswith("row_"): return jsonify({"error":"\u65e0\u6548\u7684\u5408\u540cID"}),400
    try: rn=int(cid.split("_")[1])
    except: return jsonify({"error":"\u65e0\u6548\u7684\u5408\u540cID"}),400
    t=None
    for c in read_contracts():
        if c["_row"]==rn: t=c; break
    if not t: return jsonify({"error":"\u5408\u540c\u4e0d\u5b58\u5728"}),404
    if not _can_write(t): return jsonify({"error":"\u6743\u9650\u4e0d\u8db3"}),403
    upd_cell(rn,"\u4e0b\u6b21\u62db\u91c7\u65f6\u95f4","")
    write_log("合同台账", "清空提醒", obj=str(t.get("\u7f16\u53f7","")))
    return jsonify({"message":"\u63d0\u9192\u5df2\u6e05\u9664"})

# ─── Statistics ───
def _stat_filter(cts, unit, dept, person, year, sd, ed):
    out = []
    for c in cts:
        if unit and unit not in c.get("\u6240\u5c5e\u90e8\u95e8",""): continue
        if dept and dept not in c.get("\u6240\u5c5e\u90e8\u95e8",""): continue
        if person and person not in c.get("\u7ecf\u529e\u4eba",""): continue
        cd = parse_date(c.get("\u91c7\u8d2d\u65f6\u95f4",""))
        if year and (cd is None or str(cd.year)!=year): continue
        if sd:
            s_ = parse_date(sd)
            if s_ and (cd is None or cd < s_): continue
        if ed:
            e_ = parse_date(ed)
            if e_ and (cd is None or cd > e_): continue
        out.append(c)
    return out


@app.route("/api/statistics")
@login_required
def api_statistics():
    st = request.args.get("type","by_dept")
    dim = request.args.get("dim","dept")
    # 维度拆分：按部门/项目支付情况、按部门/项目预算执行情况
    if st in ("pay_dept","pay_project","budget_dept","budget_project"):
        dim = "project" if st.endswith("project") else "dept"
        st = "unpaid" if st.startswith("pay_") else "budget"
    cts = _stat_filter(_scope_contracts(read_contracts()),
                       request.args.get("unit","").strip(),
                       request.args.get("dept","").strip(),
                       request.args.get("person","").strip(),
                       request.args.get("year","").strip(),
                       request.args.get("start_date","").strip(),
                       request.args.get("end_date","").strip())
    dim_label = "\u9879\u76ee" if dim=="project" else "\u90e8\u95e8"
    def _gkey(c, dim):
        if dim=="project":
            return c.get("\u9879\u76ee\u540d\u79f0","") or "\u672a\u77e5"
        return c.get("\u6240\u5c5e\u90e8\u95e8","") or "\u672a\u77e5"
    if st in ("unpaid","budget"):
        grp={}
        for c in cts:
            k=_gkey(c, dim)
            if k not in grp: grp[k]={"count":0,"amt":0.0,"paid":0.0,"unpaid":0.0,"bid":0.0}
            g=grp[k]; g["count"]+=1
            a=_num(c.get("\u5408\u540c\u91d1\u989d\uff08\u4e07\u5143\uff09")) or 0.0
            g["amt"]+=a
            p_=_num(c.get("\u5df2\u652f\u4ed8")) or 0.0
            g["paid"]+=p_
            up=_num(c.get("\u672a\u652f\u4ed8"))
            if up is None:
                up = (a - p_) if (a or p_) else 0.0
            g["unpaid"]+=up
            b=_num(c.get("\u51b3\u7b56\u7684\u62db\u6807\u63a7\u5236\u4ef7")) or 0.0
            g["bid"]+=b
        rows=[]
        for k,v in sorted(grp.items(), key=lambda x:-x[1]["count"]):
            if st=="unpaid":
                rows.append({dim_label:k,"\u5408\u540c\u6570":v["count"],"\u5408\u540c\u91d1\u989d":round(v["amt"],2),"\u5df2\u652f\u4ed8":round(v["paid"],2),"\u672a\u652f\u4ed8":round(v["unpaid"],2)})
            else:
                diff=v["amt"]-v["bid"]
                rate=round(diff/v["bid"]*100,2) if v["bid"] else None
                rows.append({dim_label:k,"\u5408\u540c\u6570":v["count"],"\u62db\u6807\u63a7\u5236\u4ef7":round(v["bid"],2),"\u5408\u540c\u91d1\u989d":round(v["amt"],2),"\u5dee\u989d":round(diff,2),"\u504f\u5dee\u7387":rate})
        return jsonify({"type":st,"dim":dim,"data":rows})
    if st=="by_dept":
        grp={}
        for c in cts:
            d=c.get("\u6240\u5c5e\u90e8\u95e8","\u672a\u77e5")
            if d not in grp: grp[d]={"count":0,"amt":0.0}
            grp[d]["count"]+=1
            a=_num(c.get("\u5408\u540c\u91d1\u989d\uff08\u4e07\u5143\uff09"))
            if a: grp[d]["amt"]+=a
        return jsonify({"type":"by_dept","data":[{"\u90e8\u95e8":k,"\u5408\u540c\u6570":v["count"],"\u5408\u540c\u91d1\u989d":round(v["amt"],2)} for k,v in sorted(grp.items(),key=lambda x:-x[1]["count"])]})
    elif st=="by_category":
        grp={}
        for c in cts:
            cat=c.get("\u9879\u76ee\u5206\u7c7b","\u672a\u77e5")
            if cat not in grp: grp[cat]={"count":0,"amt":0.0}
            grp[cat]["count"]+=1
            a=_num(c.get("\u5408\u540c\u91d1\u989d\uff08\u4e07\u5143\uff09"))
            if a: grp[cat]["amt"]+=a
        return jsonify({"type":"by_category","data":[{"\u5206\u7c7b":k,"\u5408\u540c\u6570":v["count"],"\u5408\u540c\u91d1\u989d":round(v["amt"],2)} for k,v in sorted(grp.items(),key=lambda x:-x[1]["count"])]})
    elif st=="by_fund_source":
        grp={}
        for c in cts:
            fs=c.get("资金来源","") or "未指定"
            if fs not in grp: grp[fs]={"count":0,"amt":0.0}
            grp[fs]["count"]+=1
            a=_num(c.get("合同金额（万元）"))
            if a: grp[fs]["amt"]+=a
        return jsonify({"type":"by_fund_source","data":[{"资金来源":k,"合同数":v["count"],"合同金额":round(v["amt"],2)} for k,v in sorted(grp.items(),key=lambda x:-x[1]["count"])]})
    elif st=="by_person":
        grp={}
        for c in cts:
            for p_ in _split_names(c.get("\u7ecf\u529e\u4eba","")):
                if p_ not in grp: grp[p_]={"count":0,"amt":0.0}
                grp[p_]["count"]+=1
                a=_num(c.get("\u5408\u540c\u91d1\u989d\uff08\u4e07\u5143\uff09"))
                if a: grp[p_]["amt"]+=a
        return jsonify({"type":"by_person","data":[{"\u7ecf\u529e\u4eba":k,"\u5408\u540c\u6570":v["count"],"\u5408\u540c\u91d1\u989d":round(v["amt"],2)} for k,v in sorted(grp.items(),key=lambda x:-x[1]["count"])]})
    elif st=="by_time":
        grp={}
        for c in cts:
            cd=parse_date(c.get("\u91c7\u8d2d\u65f6\u95f4",""))
            if not cd: continue
            q=(cd.month-1)//3+1; k=f"{cd.year}\u5e74\u7b2c{q}\u5b63\u5ea6"
            if k not in grp: grp[k]={"count":0,"amt":0.0}
            grp[k]["count"]+=1
            a=_num(c.get("\u5408\u540c\u91d1\u989d\uff08\u4e07\u5143\uff09"))
            if a: grp[k]["amt"]+=a
        return jsonify({"type":"by_time","data":[{"\u65f6\u95f4":k,"\u5408\u540c\u6570":v["count"],"\u5408\u540c\u91d1\u989d":round(v["amt"],2)} for k,v in sorted(grp.items())]})
    return jsonify({"error":"\u4e0d\u652f\u6301\u7684\u7edf\u8ba1\u7c7b\u578b"}),400

# ─── User Management ───
@app.route("/api/users",methods=["GET"])
@login_required
def api_list_users():
    u = get_cur_user()
    if u["\u89d2\u8272"] != "\u7ba1\u7406\u5458":
        return jsonify({"items":[{"username":u["\u7528\u6237\u540d"],"name":u["\u59d3\u540d"],"department":u["\u6240\u5c5e\u90e8\u95e8"],"role":u["\u89d2\u8272"],"managed_departments":u.get("\u5206\u7ba1\u90e8\u95e8","")}]})
    return jsonify({"items":[{"username":x["\u7528\u6237\u540d"],"name":x["\u59d3\u540d"],"department":x["\u6240\u5c5e\u90e8\u95e8"],"role":x["\u89d2\u8272"],"managed_departments":x.get("\u5206\u7ba1\u90e8\u95e8","")} for x in get_all_users()]})

@app.route("/api/users",methods=["POST"])
@admin_required
def api_create_user():
    d=request.get_json(silent=True) or {}
    un=d.get("username","").strip(); nm=d.get("name","").strip(); dp=d.get("department","").strip(); rl=d.get("role","普通用户").strip()
    if not un or not nm or not dp: return jsonify({"error":"用户名、姓名、所属部门不能为空"}),400
    if rl not in ("管理员","普通用户","查询用户"): return jsonify({"error":"角色无效"}),400
    if get_user(un): return jsonify({"error":"用户名已存在"}),400
    if dp not in get_dept_strings(): return jsonify({"error":"部门不存在，请先添加部门"}),400
    managed = ""
    if is_leadership_dept(dp) and d.get("managed_departments"):
        try:
            managed = norm_managed(d["managed_departments"])
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
    db_exec('INSERT INTO users ("用户名","密码","姓名","所属部门","角色","分管部门") VALUES (?,?,?,?,?,?)',
            (un, hash_pw("123456"), nm, dp, rl, managed))
    write_log("用户管理", "新增用户", obj=un, detail=nm + " / " + dp)
    return jsonify({"message":"用户创建成功，默认密码 123456"}),201

@app.route("/api/users/<username>",methods=["PUT"])
@admin_required
def api_update_user(username):
    d=request.get_json(silent=True) or {}
    cur_user = get_user(username)
    if not cur_user: return jsonify({"error":"用户不存在"}),404
    final_dept = d["department"].strip() if (d.get("department") and d["department"].strip()) else cur_user.get("所属部门","")
    sets=[]; params=[]
    if "name" in d and d["name"]:
        sets.append('"姓名"=?'); params.append(d["name"].strip())
    if "department" in d and d["department"]:
        if d["department"].strip() not in get_dept_strings(): return jsonify({"error":"部门不存在"}),400
        sets.append('"所属部门"=?'); params.append(d["department"].strip())
    if "role" in d and d["role"]:
        if d["role"] not in ("管理员","普通用户","查询用户"): return jsonify({"error":"角色无效"}),400
        sets.append('"角色"=?'); params.append(d["role"])
    if "reset_password" in d and d["reset_password"]:
        sets.append('"密码"=?'); params.append(hash_pw("123456"))
    if "managed_departments" in d:
        if is_leadership_dept(final_dept):
            try:
                managed = norm_managed(d["managed_departments"])
            except ValueError as e:
                return jsonify({"error": str(e)}), 400
            sets.append('"分管部门"=?'); params.append(managed)
        else:
            sets.append('"分管部门"=?'); params.append("")
    if sets:
        params.append(username)
        db_exec("UPDATE users SET " + ",".join(sets) + ' WHERE "用户名"=?', params)
    if "reset_password" in d and d["reset_password"]:
        write_log("用户管理", "重置密码", obj=username)
    write_log("用户管理", "修改用户", obj=username)
    return jsonify({"message":"用户更新成功"})

@app.route("/api/users/<username>/self",methods=["PUT"])
@login_required
def api_update_self(username):
    u = get_cur_user()
    if not u or u["\u7528\u6237\u540d"] != username:
        return jsonify({"error":"\u6743\u9650\u4e0d\u8db3"}),403
    cur = get_user(username)
    if not cur: return jsonify({"error":"\u7528\u6237\u4e0d\u5b58\u5728"}),404
    d = request.get_json(silent=True) or {}
    new_un = d.get("new_username","").strip()
    old_pw = str(d.get("old_password","") or "")
    new_pw = str(d.get("new_password","") or "")
    if new_pw and not old_pw:
        return jsonify({"error":"\u4fee\u6539\u5bc6\u7801\u9700\u586b\u5199\u539f\u5bc6\u7801"}),400
    if old_pw:
        stored = cur.get("\u5bc6\u7801\u54c8\u5e0c", cur.get("\u5bc6\u7801",""))
        if stored != hash_pw(old_pw):
            return jsonify({"error":"\u539f\u5bc6\u7801\u4e0d\u6b63\u786e"}),400
    if not new_un and not new_pw:
        return jsonify({"error":"\u8bf7\u586b\u5199\u8981\u4fee\u6539\u7684\u7528\u6237\u540d\u6216\u5bc6\u7801"}),400
    if new_un and new_un != username:
        if get_user(new_un):
            return jsonify({"error":"\u7528\u6237\u540d\u5df2\u5b58\u5728"}),400
        db_exec('UPDATE users SET "用户名"=? WHERE "用户名"=?', (new_un, username))
        username = new_un
    if new_pw:
        db_exec('UPDATE users SET "密码"=? WHERE "用户名"=?', (hash_pw(new_pw), username))
    session["username"] = username
    write_log("用户管理", "修改个人信息", obj=username, detail="用户名或密码变更")
    return jsonify({"message":"\u4fee\u6539\u6210\u529f","username":username})

@app.route("/api/users/<username>",methods=["DELETE"])
@admin_required
def api_delete_user(username):
    if username=="admin": return jsonify({"error":"不能删除默认管理员"}),400
    if not get_user(username): return jsonify({"error":"用户不存在"}),404
    db_exec('DELETE FROM users WHERE "用户名"=?', (username,))
    write_log("用户管理", "删除用户", obj=username)
    return jsonify({"message":"用户删除成功"})

# ─── Department Management ───
@app.route("/api/departments",methods=["GET"])
@login_required
def api_list_depts():
    return jsonify({"items":[{"unit":d["\u5355\u4f4d\u540d\u79f0"],"dept":d["\u79d1\u5ba4\u540d\u79f0"],"display":f"{d['\u5355\u4f4d\u540d\u79f0']}-{d['\u79d1\u5ba4\u540d\u79f0']}"} for d in get_all_depts()]})

@app.route("/api/departments",methods=["POST"])
@admin_required
def api_create_dept():
    d=request.get_json(silent=True) or {}; un=d.get("unit","").strip(); dn=d.get("dept","").strip()
    if not un or not dn: return jsonify({"error":"单位名称和科室名称不能为空"}),400
    for x in get_all_depts():
        if x["单位名称"]==un and x["科室名称"]==dn: return jsonify({"error":"该部门组合已存在"}),400
    db_exec('INSERT INTO departments ("单位名称","科室名称") VALUES (?,?)', (un,dn))
    write_log("部门管理", "新增部门", obj=un + "-" + dn)
    return jsonify({"message":"部门创建成功"}),201

@app.route("/api/departments",methods=["DELETE"])
@admin_required
def api_delete_dept():
    un=request.args.get("unit","").strip(); dn=request.args.get("dept","").strip()
    if not un or not dn: return jsonify({"error":"单位名称和科室名称不能为空"}),400
    ds=f"{un}-{dn}"
    for u in get_all_users():
        if u.get("所属部门","")==ds: return jsonify({"error":f"部门 '{ds}' 下仍有用户，无法删除"}),400
    depts=get_all_depts()
    if not any(x["单位名称"]==un and x["科室名称"]==dn for x in depts): return jsonify({"error":"部门不存在"}),404
    db_exec('DELETE FROM departments WHERE "单位名称"=? AND "科室名称"=?', (un,dn))
    write_log("部门管理", "删除部门", obj=un + "-" + dn)
    return jsonify({"message":"部门删除成功"})

@app.route("/api/departments",methods=["PUT"])
@admin_required
def api_update_dept():
    d=request.get_json(silent=True) or {}
    un=d.get("unit","").strip(); dn=d.get("dept","").strip()
    n_un=d.get("new_unit","").strip(); n_dn=d.get("new_dept","").strip()
    if not un or not dn or not n_un or not n_dn:
        return jsonify({"error":"单位名称和部门名称不能为空"}),400
    old=f"{un}-{dn}"; new=f"{n_un}-{n_dn}"
    depts=get_all_depts()
    if not any(x["单位名称"]==un and x["科室名称"]==dn for x in depts):
        return jsonify({"error":"部门不存在"}),404
    if old!=new and any(x["单位名称"]==n_un and x["科室名称"]==n_dn for x in depts):
        return jsonify({"error":"该部门组合已存在"}),400
    db_exec('UPDATE departments SET "单位名称"=?, "科室名称"=? WHERE "单位名称"=? AND "科室名称"=?',(n_un,n_dn,un,dn))
    if old!=new:
        db_exec('UPDATE users SET "所属部门"=? WHERE "所属部门"=?',(new,old))
        for u in get_all_users():
            mg=str(u.get("分管部门") or "").strip()
            if not mg: continue
            parts=[p.strip() for p in mg.replace("、",",").replace("；",";").split(",") if p.strip()]
            if old in parts:
                parts=[new if p==old else p for p in parts]
                db_exec('UPDATE users SET "分管部门"=? WHERE "用户名"=?',("、".join(parts),u["用户名"]))
        db_exec('UPDATE contracts SET "所属部门"=? WHERE "所属部门"=?',(new,old))
    write_log("部门管理","编辑部门",obj=old+" → "+new)
    return jsonify({"message":"部门修改成功"})

# ─── Config ───
@app.route("/api/config",methods=["GET"])
@login_required
def api_get_config():
    return jsonify({"default_remind_days":get_default_days(),"log_keep_days":get_log_keep_days(),"default_attach_max_mb":get_attach_max_mb()})

@app.route("/api/config",methods=["PUT"])
@admin_required
def api_update_config():
    d=request.get_json(silent=True) or {}
    updated=[]
    if "default_remind_days" in d:
        try:
            v=int(d["default_remind_days"])
            if v<1 or v>365: return jsonify({"error":"\u63d0\u9192\u5929\u6570\u5e94\u57281-365\u4e4b\u95f4"}),400
            set_cfg("default_remind_days",v)
            write_log("系统配置", "修改配置", obj="default_remind_days", detail=str(v))
            updated.append("default_remind_days")
        except: return jsonify({"error":"\u65e0\u6548\u7684\u6570\u503c"}),400
    if "log_keep_days" in d:
        try:
            v=int(d["log_keep_days"])
            if v<1 or v>3650: return jsonify({"error":"\u65e5\u5fd7\u4fdd\u7559\u5929\u6570\u5e94\u57281-3650\u4e4b\u95f4"}),400
            set_cfg("log_keep_days",v)
            write_log("系统配置", "修改配置", obj="log_keep_days", detail=str(v))
            updated.append("log_keep_days")
        except: return jsonify({"error":"\u65e0\u6548\u7684\u6570\u503c"}),400
    if "default_attach_max_mb" in d:
        try:
            v=int(d["default_attach_max_mb"])
            if v<1 or v>200: return jsonify({"error":"附件大小限制应在1-200(M)之间"}),400
            set_cfg("default_attach_max_mb",v)
            write_log("系统配置", "修改配置", obj="default_attach_max_mb", detail=str(v))
            updated.append("default_attach_max_mb")
        except: return jsonify({"error":"无效的数值"}),400
    if not updated: return jsonify({"error":"\u7f3a\u5c11\u914d\u7f6e\u9879"}),400
    return jsonify({"message":"\u914d\u7f6e\u66f4\u65b0\u6210\u529f","default_remind_days":get_default_days(),"log_keep_days":get_log_keep_days(),"default_attach_max_mb":get_attach_max_mb()})

# ─── 操作日志（审计） ───
def _log_filter():
    # 分级查看：管理员可看全量；普通/查询用户仅看本人
    u = get_cur_user()
    a = request.args
    start_date = a.get("start_date","").strip()
    end_date = a.get("end_date","").strip()
    username = a.get("username","").strip()
    module = a.get("module","").strip()
    action = a.get("action","").strip()
    conds=[]; params=[]
    if u and u["\u89d2\u8272"] != "\u7ba1\u7406\u5458":
        conds.append('"用户名" = ?'); params.append(u["\u7528\u6237\u540d"])
    elif username:
        conds.append('"用户名" LIKE ?'); params.append("%" + username + "%")
    if start_date: conds.append('"时间" >= ?'); params.append(start_date + " 00:00:00")
    if end_date: conds.append('"时间" <= ?'); params.append(end_date + " 23:59:59")
    if module: conds.append('"模块" = ?'); params.append(module)
    if action: conds.append('"操作类型" = ?'); params.append(action)
    where = (" WHERE " + " AND ".join(conds)) if conds else ""
    return where, params

@app.route("/api/logs")
@login_required
def api_list_logs():
    where, params = _log_filter()
    a = request.args
    try:
        page = max(1, int(a.get("page", 1))); ps = min(200, max(1, int(a.get("page_size", 20))))
    except Exception:
        page = 1; ps = 20
    total = db_query('SELECT COUNT(*) c FROM operation_logs' + where, params)[0]["c"]
    rows = db_query('SELECT * FROM operation_logs' + where + ' ORDER BY id DESC LIMIT ? OFFSET ?',
                    params + [ps, (page-1)*ps])
    return jsonify({"total": total, "page": page, "page_size": ps, "items": rows})

@app.route("/api/logs/export")
@admin_required
def api_export_logs():
    where, params = _log_filter()
    rows = db_query('SELECT * FROM operation_logs' + where + ' ORDER BY id DESC', params)
    headers = ["时间","用户名","姓名","角色","IP地址","模块","操作类型","对象","详情","结果"]
    header_keys = {"IP地址": "IP"}
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "操作日志"
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    for ri, r in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=r.get(header_keys.get(h, h), ""))
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    ts = datetime.datetime.now().strftime("%Y%m%d")
    from urllib.parse import quote
    fn = f"操作日志_export_{ts}.xlsx"
    write_log("日志审计", "导出", obj="操作日志", detail="导出 " + str(len(rows)) + " 条")
    return Response(bio.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fn)}"})

# âââ Root Route âââ
@app.route("/")
def index():
    idx_path = os.path.join(BASE_DIR, "templates", "index.html")
    if os.path.exists(idx_path):
        return send_from_directory(os.path.join(BASE_DIR, "templates"), "index.html")
    html = '<!DOCTYPE html><html lang="zh-CN"><head><meta charset="utf-8"><title>潭合同 台账管理系统</title><style>body{font-family:sans-serif;display:flex;justify-content:center;align-items:center;min-height:100vh;margin:0;background:linear-gradient(135deg,#0A2E5C,#1E90FF);color:#fff}.card{text-align:center;padding:40px;background:rgba(255,255,255,0.1);border-radius:16px;backdrop-filter:blur(10px)}h1{font-size:28px;margin:0 0 8px}p{font-size:14px;opacity:.8}</style></head><body><div class="card"><h1>潭合同</h1><p>台账管理系统</p><p>湘潭市机关事务管理局</p><hr style="width:40px;margin:16px auto;border-color:rgba(255,255,255,.3)"><p style="font-size:12px;opacity:.6">请将前端 index.html 放置于 templates/ 目录</p></div></body></html>'
    return html

# âââ Error Handlers âââ
@app.errorhandler(404)
def not_found(e):
    if request.path.startswith("/api/"):
        return jsonify({"error":"\u63a5\u53e3\u4e0d\u5b58\u5728"}),404
    return index()

@app.errorhandler(500)
def server_error(e): return jsonify({"error":"\u670d\u52a1\u5668\u5185\u90e8\u9519\u8bef"}),500


# ═══════════════════════════════════════════════════════════════════════
init_db()
_maybe_cleanup_logs()

if __name__=="__main__":
    lp=os.path.join(STATIC_DIR,"logo.jpg")
    if not os.path.exists(lp): print("[WARN] static/logo.jpg 不存在，请将单位LOGO文件放置于此路径。")
    print("[INFO] 潭合同 台账管理系统 启动中...")
    print(f"[INFO] 数据库: {DB_PATH}")
    print("[INFO] 访问地址: http://localhost:5000")
    print("[INFO] 默认管理员: admin / 123456")
    import webbrowser; webbrowser.open("http://localhost:5000")
    app.run(host="0.0.0.0",port=5000,debug=False,threaded=True)
